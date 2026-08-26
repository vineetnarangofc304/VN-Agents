"""
LinkedIn CRM — Campaign & Connections Manager
User-isolated campaigns, XLSX upload, smart send (message + connection request fallback).
"""
import os
import io
import uuid
import random
import string
import logging
import asyncio
import openpyxl
from datetime import datetime, timezone
from fastapi import APIRouter, HTTPException, Request, UploadFile, File, Form
from motor.motor_asyncio import AsyncIOMotorClient
import httpx

from routes.crm_auth import get_current_user

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/crm", tags=["crm"])

mongo_url = os.environ.get("MONGO_URL", "")
_client = AsyncIOMotorClient(mongo_url)
_db = _client[os.environ.get("DB_NAME", "test_database")]

MAX_DAILY_SENDS = 50


# ============ Helpers ============
def _build_headers(li_at: str, jsessionid: str) -> dict:
    clean_js = jsessionid.strip('"').replace("ajax:", "")
    return {
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "accept": "application/vnd.linkedin.normalized+json+2.1",
        "x-restli-protocol-version": "2.0.0",
        "content-type": "application/json; charset=UTF-8",
        "cookie": f'li_at={li_at}; JSESSIONID="ajax:{clean_js}"',
        "csrf-token": f"ajax:{clean_js}",
    }


async def _get_user_cookies(user: dict) -> tuple:
    """Get user's LinkedIn cookies. Returns (li_at, jsessionid) or raises."""
    full = await _db.crm_users.find_one({"email": user["email"]})
    li_at = (full or {}).get("li_at", "")
    jsessionid = (full or {}).get("jsessionid", "")
    if not li_at or not jsessionid:
        raise HTTPException(status_code=400, detail="LinkedIn cookies not configured. Go to Settings and add your li_at + JSESSIONID.")
    return li_at, jsessionid


async def _resolve_urn(http_client, public_id: str, headers: dict):
    """Resolve LinkedIn public_id to fsd_profile URN. Falls back gracefully."""
    try:
        resp = await http_client.get(
            f"https://www.linkedin.com/voyager/api/identity/dash/profiles?q=memberIdentity&memberIdentity={public_id}",
            headers=headers,
        )
        if resp.status_code == 200:
            for item in resp.json().get("included", []):
                eid = item.get("entityUrn", "")
                if "fsd_profile" in eid:
                    return eid, None
    except Exception:
        pass
    # Return None — caller will use public_id fallback
    return None, None


async def _send_message_or_invite(http_client, headers: dict, recipient_urn: str, public_id: str, direct_msg: str, invite_note: str):
    """Try direct message first, fall back to connection request.
    Uses URN if available, otherwise uses public_id for invite."""

    # Try direct message if we have a URN (only works for 1st-degree)
    if recipient_urn:
        member_id = recipient_urn.split(":")[-1] if ":" in recipient_urn else recipient_urn
        tracking = "".join(random.choices(string.ascii_lowercase + string.digits, k=16))

        payload = {
            "keyVersion": "LEGACY_INBOX",
            "conversationCreate": {
                "eventCreate": {
                    "originToken": str(uuid.uuid4()),
                    "value": {
                        "com.linkedin.voyager.messaging.create.MessageCreate": {
                            "body": direct_msg,
                            "attributedBody": {"text": direct_msg, "attributes": []},
                            "attachments": [],
                        }
                    },
                    "trackingId": tracking,
                },
                "dedupeByClientGeneratedToken": False,
                "recipients": [member_id],
                "subtype": "MEMBER_TO_MEMBER",
            }
        }
        resp = await http_client.post(
            "https://www.linkedin.com/voyager/api/messaging/conversations?action=create",
            json=payload, headers=headers,
        )
        if resp.status_code in [200, 201]:
            return True, "message_sent"

    # Connection request — try verifyQuotaAndCreate first (with URN or public_id)
    for attempt_urn in [recipient_urn, f"urn:li:fsd_profile:{public_id}" if not recipient_urn else None]:
        if not attempt_urn:
            continue
        invite_payload = {
            "inviteeProfileUrn": attempt_urn,
            "customMessage": invite_note[:300],
            "trackingId": str(uuid.uuid4()),
        }
        resp2 = await http_client.post(
            "https://www.linkedin.com/voyager/api/voyagerRelationshipsDashMemberRelationships?action=verifyQuotaAndCreate",
            json=invite_payload, headers=headers,
        )
        if resp2.status_code in [200, 201]:
            return True, "invite_sent"
        if resp2.status_code == 401:
            return False, "cookie_expired_401"

    # Fallback: normInvitations with public_id
    invite_payload_v2 = {
        "trackingId": str(uuid.uuid4()),
        "message": invite_note[:300],
        "invitations": [],
        "excludeInvitations": [],
        "invitee": {
            "com.linkedin.voyager.growth.invitation.InviteeProfile": {
                "profileId": public_id
            }
        }
    }
    resp3 = await http_client.post(
        "https://www.linkedin.com/voyager/api/growth/normInvitations",
        json=invite_payload_v2, headers=headers,
    )
    if resp3.status_code in [200, 201]:
        return True, "invite_sent"
    if resp3.status_code == 401:
        return False, "cookie_expired_401"

    return False, f"send_failed_{resp3.status_code}"


# ============ Campaigns ============
@router.get("/campaigns")
async def list_campaigns(request: Request):
    user = await get_current_user(request)
    campaigns = []
    async for c in _db.crm_campaigns.find({"user_id": user["id"]}, {"_id": 0}).sort("created_at", -1):
        cid = c["campaign_id"]
        total = await _db.crm_prospects.count_documents({"campaign_id": cid})
        sent = await _db.crm_prospects.count_documents({"campaign_id": cid, "status": "sent"})
        failed = await _db.crm_prospects.count_documents({"campaign_id": cid, "status": "failed"})
        pending = await _db.crm_prospects.count_documents({"campaign_id": cid, "status": "pending"})
        c.update({"total": total, "sent": sent, "failed": failed, "pending": pending})
        campaigns.append(c)
    return {"campaigns": campaigns}


@router.post("/campaigns")
async def create_campaign(request: Request):
    user = await get_current_user(request)
    data = await request.json()
    name = data.get("name", "Untitled Campaign")
    direct_message = data.get("direct_message", "")
    invite_note = data.get("invite_note", "")
    daily_limit = min(int(data.get("daily_limit", 25)), MAX_DAILY_SENDS)

    if not direct_message and not invite_note:
        raise HTTPException(status_code=400, detail="At least one message template required")

    campaign_id = str(uuid.uuid4())[:8]
    doc = {
        "campaign_id": campaign_id,
        "user_id": user["id"],
        "user_name": user["name"],
        "name": name,
        "direct_message": direct_message,
        "invite_note": invite_note,
        "daily_limit": daily_limit,
        "status": "active",
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await _db.crm_campaigns.insert_one(doc)
    return {"success": True, "campaign_id": campaign_id}


@router.get("/campaigns/{campaign_id}")
async def get_campaign(campaign_id: str, request: Request):
    user = await get_current_user(request)
    campaign = await _db.crm_campaigns.find_one({"campaign_id": campaign_id, "user_id": user["id"]}, {"_id": 0})
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    # Get prospects
    prospects = []
    async for p in _db.crm_prospects.find({"campaign_id": campaign_id}, {"_id": 0}).sort("rank", 1):
        prospects.append(p)
    campaign["prospects"] = prospects
    return campaign


@router.post("/campaigns/{campaign_id}/upload")
async def upload_prospects(campaign_id: str, request: Request, file: UploadFile = File(...)):
    user = await get_current_user(request)
    campaign = await _db.crm_campaigns.find_one({"campaign_id": campaign_id, "user_id": user["id"]})
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content))
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid or corrupt XLSX file. Please upload a valid spreadsheet.")
    ws = wb.active

    # Find header row (look for "name" or "contact" in first 10 rows)
    header_row = 1
    headers_map = {}
    for row_num in range(1, min(11, ws.max_row + 1)):
        row_vals = [str(ws.cell(row=row_num, column=c).value or "").strip().lower() for c in range(1, ws.max_column + 1)]
        has_name = any("name" in v for v in row_vals)
        has_linkedin = any("linkedin" in v for v in row_vals)
        if has_name or has_linkedin:
            header_row = row_num
            for i, v in enumerate(row_vals):
                headers_map[v] = i
            break

    # Find column indices
    name_col = next((v for k, v in headers_map.items() if "contact name" in k or k == "name" or "full name" in k), None)
    linkedin_col = next((v for k, v in headers_map.items() if "linkedin" in k), None)
    company_col = next((v for k, v in headers_map.items() if "company" in k), None)
    title_col = next((v for k, v in headers_map.items() if "title" in k or "role" in k), None)

    if linkedin_col is None:
        raise HTTPException(status_code=400, detail="No 'LinkedIn' column found in the spreadsheet")

    imported = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        row = list(row)
        if not row or len(row) <= linkedin_col:
            continue
        linkedin = str(row[linkedin_col] or "").strip()
        if "linkedin.com" not in linkedin:
            continue
        name = str(row[name_col] or "").strip() if name_col is not None and len(row) > name_col else ""
        company = str(row[company_col] or "").strip() if company_col is not None and len(row) > company_col else ""
        title = str(row[title_col] or "").strip() if title_col is not None and len(row) > title_col else ""
        public_id = linkedin.rstrip("/").split("/")[-1].split("?")[0]

        if not public_id:
            continue

        existing = await _db.crm_prospects.find_one({"campaign_id": campaign_id, "public_id": public_id})
        if existing:
            continue

        await _db.crm_prospects.insert_one({
            "campaign_id": campaign_id,
            "rank": imported + 1,
            "name": name,
            "company": company,
            "title": title,
            "linkedin_url": linkedin,
            "public_id": public_id,
            "status": "pending",
            "send_type": None,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "sent_at": None,
            "error": None,
        })
        imported += 1

    return {"success": True, "imported": imported, "campaign_id": campaign_id}


@router.post("/campaigns/{campaign_id}/send")
async def send_campaign_batch(campaign_id: str, request: Request, count: int = 10):
    user = await get_current_user(request)
    campaign = await _db.crm_campaigns.find_one({"campaign_id": campaign_id, "user_id": user["id"]})
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")

    li_at, jsessionid = await _get_user_cookies(user)
    headers = _build_headers(li_at, jsessionid)
    count = min(count, campaign.get("daily_limit", 25))

    # Get pending prospects
    prospects = []
    async for p in _db.crm_prospects.find({"campaign_id": campaign_id, "status": "pending"}).sort("rank", 1).limit(count):
        prospects.append(p)

    if not prospects:
        return {"success": True, "detail": "No pending prospects", "sent": 0, "failed": 0}

    # Send in background
    asyncio.create_task(_send_batch_bg(campaign, prospects, headers))

    return {
        "success": True,
        "detail": f"Sending {len(prospects)} in background",
        "batch_size": len(prospects),
    }


async def _send_batch_bg(campaign: dict, prospects: list, headers: dict):
    """Background task to send messages."""
    sent_count = 0
    failed_count = 0
    direct_msg = campaign.get("direct_message", "")
    invite_note = campaign.get("invite_note", "")

    async with httpx.AsyncClient(timeout=20.0) as client:
        for prospect in prospects:
            try:
                # Check if campaign was stopped
                camp_check = await _db.crm_campaigns.find_one({"campaign_id": campaign["campaign_id"]})
                if camp_check and camp_check.get("status") == "paused":
                    logger.info(f"CRM Campaign {campaign['campaign_id']}: Stopped by user")
                    break

                # Mark as sending
                await _db.crm_prospects.update_one(
                    {"campaign_id": campaign["campaign_id"], "public_id": prospect["public_id"]},
                    {"$set": {"status": "sending"}}
                )

                # Resolve URN (optional — we can send invites without it)
                urn, _ = await _resolve_urn(client, prospect["public_id"], headers)

                # Personalize messages
                first_name = prospect.get("name", "").split()[0] if prospect.get("name") else "there"
                company = prospect.get("company") or "your company"
                title = prospect.get("title") or ""
                p_direct = direct_msg.replace("{name}", first_name).replace("{company}", company).replace("{title}", title)
                p_invite = invite_note.replace("{name}", first_name).replace("{company}", company).replace("{title}", title)

                # Send — passes both URN and public_id so it can fallback
                success, send_type = await _send_message_or_invite(client, headers, urn, prospect["public_id"], p_direct, p_invite)

                if success:
                    await _db.crm_prospects.update_one(
                        {"campaign_id": campaign["campaign_id"], "public_id": prospect["public_id"]},
                        {"$set": {"status": "sent", "send_type": send_type, "sent_at": datetime.now(timezone.utc).isoformat()}}
                    )
                    sent_count += 1
                    logger.info(f"CRM Campaign {campaign['campaign_id']}: {send_type} to {prospect['name']} ({prospect['public_id']})")
                else:
                    await _db.crm_prospects.update_one(
                        {"campaign_id": campaign["campaign_id"], "public_id": prospect["public_id"]},
                        {"$set": {"status": "failed", "error": send_type, "sent_at": datetime.now(timezone.utc).isoformat()}}
                    )
                    failed_count += 1
                    # If we get a rate limit / auth error, stop the batch
                    if "500" in str(send_type) or "429" in str(send_type) or "401" in str(send_type) or "cookie" in str(send_type).lower():
                        logger.warning(f"CRM Campaign {campaign['campaign_id']}: Auth/rate error ({send_type}), stopping batch")
                        break

                # Human-like delay: 25-35 seconds between sends
                delay = random.randint(25, 35)
                await asyncio.sleep(delay)

            except Exception as e:
                logger.error(f"CRM send error for {prospect.get('name')}: {e}")
                await _db.crm_prospects.update_one(
                    {"campaign_id": campaign["campaign_id"], "public_id": prospect["public_id"]},
                    {"$set": {"status": "failed", "error": str(e)[:200]}}
                )
                failed_count += 1

    logger.info(f"CRM Campaign {campaign['campaign_id']}: Batch done. Sent={sent_count}, Failed={failed_count}")


@router.put("/campaigns/{campaign_id}")
async def update_campaign(campaign_id: str, request: Request):
    user = await get_current_user(request)
    campaign = await _db.crm_campaigns.find_one({"campaign_id": campaign_id, "user_id": user["id"]})
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")

    data = await request.json()
    fields = {}
    for k in ["name", "direct_message", "invite_note", "daily_limit", "status"]:
        if k in data:
            fields[k] = data[k]
    if "daily_limit" in fields:
        fields["daily_limit"] = min(int(fields["daily_limit"]), MAX_DAILY_SENDS)

    if not fields:
        raise HTTPException(status_code=400, detail="No fields to update")

    await _db.crm_campaigns.update_one({"campaign_id": campaign_id}, {"$set": fields})
    return {"success": True, "detail": "Campaign updated"}


@router.post("/campaigns/{campaign_id}/stop")
async def stop_campaign(campaign_id: str, request: Request):
    user = await get_current_user(request)
    campaign = await _db.crm_campaigns.find_one({"campaign_id": campaign_id, "user_id": user["id"]})
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    await _db.crm_campaigns.update_one({"campaign_id": campaign_id}, {"$set": {"status": "paused"}})
    # Also revert any "sending" back to "pending"
    await _db.crm_prospects.update_many(
        {"campaign_id": campaign_id, "status": "sending"},
        {"$set": {"status": "pending"}}
    )
    return {"success": True, "detail": "Campaign stopped"}


@router.delete("/campaigns/{campaign_id}")
async def delete_campaign(campaign_id: str, request: Request):
    user = await get_current_user(request)
    campaign = await _db.crm_campaigns.find_one({"campaign_id": campaign_id, "user_id": user["id"]})
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    await _db.crm_campaigns.delete_one({"campaign_id": campaign_id})
    await _db.crm_prospects.delete_many({"campaign_id": campaign_id})
    return {"success": True, "detail": "Campaign and all prospects deleted"}


@router.post("/campaigns/{campaign_id}/retry")
async def retry_failed(campaign_id: str, request: Request):
    user = await get_current_user(request)
    campaign = await _db.crm_campaigns.find_one({"campaign_id": campaign_id, "user_id": user["id"]})
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    result = await _db.crm_prospects.update_many(
        {"campaign_id": campaign_id, "status": "failed"},
        {"$set": {"status": "pending", "error": None, "sent_at": None, "send_type": None}}
    )
    return {"success": True, "reset": result.modified_count}


@router.post("/campaigns/{campaign_id}/prospect-status")
async def update_prospect_status(campaign_id: str, request: Request):
    """Update a single prospect's status — used by the local sender script."""
    user = await get_current_user(request)
    campaign = await _db.crm_campaigns.find_one({"campaign_id": campaign_id, "user_id": user["id"]})
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")

    data = await request.json()
    public_id = data.get("public_id", "")
    status = data.get("status", "")
    send_type = data.get("send_type")
    error = data.get("error")

    if not public_id or status not in ["sent", "failed", "pending"]:
        raise HTTPException(status_code=400, detail="public_id and valid status required")

    fields = {"status": status, "sent_at": datetime.now(timezone.utc).isoformat()}
    if send_type:
        fields["send_type"] = send_type
    if error:
        fields["error"] = error
    if status == "pending":
        fields["sent_at"] = None
        fields["error"] = None
        fields["send_type"] = None

    result = await _db.crm_prospects.update_one(
        {"campaign_id": campaign_id, "public_id": public_id},
        {"$set": fields}
    )
    return {"success": result.modified_count > 0}


# ============ Connections ============
@router.get("/connections")
async def get_connections(request: Request, count: int = 50, start: int = 0):
    user = await get_current_user(request)
    li_at, jsessionid = await _get_user_cookies(user)
    headers = _build_headers(li_at, jsessionid)

    async with httpx.AsyncClient(timeout=20.0) as client:
        url = f"https://www.linkedin.com/voyager/api/relationships/dash/connections?q=search&sortType=RECENTLY_ADDED&count={count}&start={start}&decorationId=com.linkedin.voyager.dash.deco.web.mynetwork.ConnectionListWithProfile-15"
        resp = await client.get(url, headers=headers)

        if resp.status_code != 200:
            raise HTTPException(status_code=502, detail=f"Failed to fetch connections (status {resp.status_code})")

        data = resp.json()
        included = data.get("included", [])

        # Build a lookup of all included entities by entityUrn
        entity_map = {}
        for item in included:
            urn = item.get("entityUrn", "")
            if urn:
                entity_map[urn] = item

        connections = []
        for item in included:
            fn = item.get("firstName", "")
            ln = item.get("lastName", "")
            if not fn or not ln:
                continue

            occupation = item.get("occupation", "") or item.get("headline", "") or ""
            public_id = item.get("publicIdentifier", "")
            urn = item.get("entityUrn", "")

            # Parse company from occupation ("Title at Company")
            company = ""
            title = occupation
            if " at " in occupation:
                parts = occupation.split(" at ", 1)
                title = parts[0].strip()
                company = parts[1].strip()
            elif " @ " in occupation:
                parts = occupation.split(" @ ", 1)
                title = parts[0].strip()
                company = parts[1].strip()

            conn = {
                "name": f"{fn} {ln}",
                "firstName": fn,
                "lastName": ln,
                "publicIdentifier": public_id,
                "headline": occupation,
                "title": title,
                "company": company,
                "urn": urn,
                "linkedinUrl": f"https://linkedin.com/in/{public_id}" if public_id else "",
                "email": "",
                "phone": "",
                "location": item.get("locationName", "") or "",
            }
            connections.append(conn)

        # Enrich with contact info (email/phone) — limit to 10 to avoid rate limits
        import asyncio as _aio
        for conn in connections[:10]:
            if not conn["publicIdentifier"]:
                continue
            try:
                contact_url = f"https://www.linkedin.com/voyager/api/identity/profiles/{conn['publicIdentifier']}/profileContactInfo"
                cr = await client.get(contact_url, headers=headers)
                if cr.status_code == 200:
                    cdata = cr.json()
                    # Email
                    emails = cdata.get("emailAddress", "") or ""
                    if not emails:
                        email_list = cdata.get("emailAddresses", []) or []
                        if email_list:
                            emails = email_list[0] if isinstance(email_list[0], str) else email_list[0].get("emailAddress", "")
                    conn["email"] = emails

                    # Phone
                    phones = cdata.get("phoneNumbers", []) or []
                    if phones:
                        conn["phone"] = phones[0].get("number", "") if isinstance(phones[0], dict) else str(phones[0])

                    # Website / Twitter
                    websites = cdata.get("websites", []) or []
                    if websites:
                        conn["website"] = websites[0].get("url", "") if isinstance(websites[0], dict) else str(websites[0])
            except Exception:
                pass  # Contact info fetch is best-effort
            await _aio.sleep(1)  # Small delay between enrichment calls

    total = data.get("paging", {}).get("total", len(connections))
    return {"connections": connections, "total": total, "start": start}


@router.post("/connections/message")
async def message_connections(request: Request):
    """Send a message to selected connections."""
    user = await get_current_user(request)
    data = await request.json()
    recipients = data.get("recipients", [])  # list of {publicId, name, urn}
    message = data.get("message", "")

    if not recipients:
        raise HTTPException(status_code=400, detail="No recipients")
    if not message:
        raise HTTPException(status_code=400, detail="Message required")

    li_at, jsessionid = await _get_user_cookies(user)
    headers = _build_headers(li_at, jsessionid)

    results = []
    async with httpx.AsyncClient(timeout=20.0) as client:
        for r in recipients[:50]:  # max 50 per batch
            member_id = r.get("urn", "").split(":")[-1] if r.get("urn") else ""
            name = r.get("name", "").split()[0] if r.get("name") else "there"
            company = r.get("company", "your company") or "your company"
            title = r.get("title", "") or ""
            personal_msg = message.replace("{name}", name).replace("{company}", company).replace("{title}", title)

            if not member_id:
                results.append({"name": r.get("name"), "status": "skipped", "error": "No URN"})
                continue

            tracking = "".join(random.choices(string.ascii_lowercase + string.digits, k=16))
            payload = {
                "keyVersion": "LEGACY_INBOX",
                "conversationCreate": {
                    "eventCreate": {
                        "originToken": str(uuid.uuid4()),
                        "value": {
                            "com.linkedin.voyager.messaging.create.MessageCreate": {
                                "body": personal_msg,
                                "attributedBody": {"text": personal_msg, "attributes": []},
                                "attachments": [],
                            }
                        },
                        "trackingId": tracking,
                    },
                    "dedupeByClientGeneratedToken": False,
                    "recipients": [member_id],
                    "subtype": "MEMBER_TO_MEMBER",
                }
            }
            resp = await client.post(
                "https://www.linkedin.com/voyager/api/messaging/conversations?action=create",
                json=payload, headers=headers,
            )
            success = resp.status_code in [200, 201]
            results.append({"name": r.get("name"), "status": "sent" if success else "failed"})

            if success:
                logger.info(f"CRM message sent to {r.get('name')} by {user['name']}")

            await asyncio.sleep(5)

    sent = sum(1 for r in results if r["status"] == "sent")
    return {"success": True, "total": len(results), "sent": sent, "results": results}


# ============ User Settings ============
@router.get("/sender-script")
async def download_sender_script():
    """Download the local LinkedIn sender script."""
    from fastapi.responses import FileResponse
    script_path = os.path.join(os.path.dirname(__file__), "..", "static", "linkedin_sender.py")
    return FileResponse(script_path, filename="linkedin_sender.py", media_type="text/x-python")


@router.get("/settings")
async def get_settings(request: Request):
    user = await get_current_user(request)
    full = await _db.crm_users.find_one({"email": user["email"]})
    if not full:
        raise HTTPException(status_code=404, detail="User not found")
    return {
        "name": full.get("name", ""),
        "email": full.get("email", ""),
        "has_cookie": bool(full.get("li_at")),
        "li_at_preview": (full.get("li_at", "") or "")[:20] + "..." if full.get("li_at") else "",
        "jsessionid_preview": (full.get("jsessionid", "") or "")[:20] + "..." if full.get("jsessionid") else "",
    }


@router.post("/settings/cookie")
async def update_cookie(request: Request):
    user = await get_current_user(request)
    data = await request.json()
    li_at = (data.get("li_at") or "").strip()
    jsessionid = (data.get("jsessionid") or "").strip()
    if not li_at:
        raise HTTPException(status_code=400, detail="li_at required")
    if not jsessionid:
        raise HTTPException(status_code=400, detail="JSESSIONID required")

    await _db.crm_users.update_one(
        {"email": user["email"]},
        {"$set": {"li_at": li_at, "jsessionid": jsessionid}}
    )
    return {"success": True, "detail": "LinkedIn cookies updated"}
