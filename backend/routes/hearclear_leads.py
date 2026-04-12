import os
import re
import uuid
import logging
import threading
import asyncio
import io
import time
import tempfile
import traceback
from pathlib import Path
from datetime import datetime, timezone
from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional, List
from motor.motor_asyncio import AsyncIOMotorClient
from pymongo import MongoClient
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/hearclear-leads", tags=["hearclear-leads"])

mongo_url = os.environ['MONGO_URL']
db_name = os.environ['DB_NAME']
client = AsyncIOMotorClient(mongo_url)
db = client[db_name]

DOWNLOADS_DIR = Path(__file__).parent.parent / "uploads" / "hearclear_leads"
DOWNLOADS_DIR.mkdir(parents=True, exist_ok=True)

# Indian phone regex: 10 digits starting with 6-9, with optional +91/0 prefix
PHONE_REGEX = re.compile(
    r'(?:(?:\+91[\s\-]?)|(?:0))?'
    r'([6-9]\d{9})'
)

CITIES = ["Delhi", "Noida", "Gurgaon", "Gurugram", "Faridabad", "Ghaziabad", "Greater Noida", "NCR"]

CATEGORIES = [
    {
        "id": "rwa",
        "name": "RWA / Resident Welfare Associations",
        "keywords": [
            "RWA directory", "RWA members list", "resident welfare association",
            "RWA president contact", "RWA office bearers", "housing society directory",
            "residents association members", "colony RWA list", "apartment owners association",
            "cooperative society members list", "RWA phone numbers"
        ]
    },
    {
        "id": "clubs",
        "name": "Club & Social Organizations",
        "keywords": [
            "club members directory", "club members list", "golf club members",
            "rotary club directory", "lions club members", "gymkhana club list",
            "social club members", "recreation club directory", "sports club members list"
        ]
    },
    {
        "id": "schools",
        "name": "Schools & Teachers",
        "keywords": [
            "school teachers list", "teachers directory", "school staff list",
            "principal contact number", "education department employees",
            "school employees directory", "college faculty list", "university staff directory",
            "teachers phone numbers", "school directory"
        ]
    },
    {
        "id": "govt",
        "name": "Government & PSU Employees",
        "keywords": [
            "government employees list", "government officers directory",
            "PSU employees", "electricity board employees", "municipal corporation staff",
            "government department directory", "civil services list", "IAS officers",
            "state government employees", "government officials phone",
            "DDA officials", "NDMC directory", "MCD officials list"
        ]
    },
    {
        "id": "professional",
        "name": "Professional Directories",
        "keywords": [
            "architects directory", "chartered accountants list", "lawyers directory",
            "doctors list", "advocates directory", "bar council members",
            "engineers association members", "professionals directory phone",
            "consultants directory", "judges list district court"
        ]
    },
    {
        "id": "religious",
        "name": "Religious & Community Organizations",
        "keywords": [
            "gurdwara committee members", "temple trust members", "masjid committee",
            "church parish directory", "religious organization directory",
            "community leaders list", "sikh sangat directory", "mandir committee members",
            "dharamshala committee", "community association members"
        ]
    },
    {
        "id": "business",
        "name": "Business & Trade Associations",
        "keywords": [
            "traders association members", "market association directory",
            "chamber of commerce members", "business directory", "shopkeepers association",
            "industrial association members", "FICCI members directory",
            "manufacturer directory", "vendor list with contact"
        ]
    },
    {
        "id": "senior",
        "name": "Senior Citizens & Elder Care",
        "keywords": [
            "senior citizens association", "senior citizens directory",
            "retired employees list", "pensioners association", "elderly care directory",
            "old age home directory", "senior citizen club members",
            "retired government officers", "ex-servicemen directory"
        ]
    }
]

FILE_TYPES = ["pdf", "xls", "xlsx", "csv"]

active_crawl = {}

# Seed URLs — verified data sources with phone numbers
SEED_URLS = [
    # RWA Directories
    {"url": "https://www.serwa.org.in/SERWA_Telephone_Directory.pdf", "category": "rwa", "category_name": "RWA / Resident Welfare Associations", "city": "Delhi"},
    {"url": "http://www.rwa17b.com/telephone-directory.html", "category": "rwa", "category_name": "RWA / Resident Welfare Associations", "city": "Delhi"},
    {"url": "https://ro.scribd.com/doc/258343370/Resident-Welfare-Society-Committe-Delhi-Contacts", "category": "rwa", "category_name": "RWA / Resident Welfare Associations", "city": "Delhi"},
    {"url": "https://www.rwadblocksaket.com/contactus.htm", "category": "rwa", "category_name": "RWA / Resident Welfare Associations", "city": "Delhi"},
    {"url": "https://www.scribd.com/document/697675425/Rwa-team-2022-24", "category": "rwa", "category_name": "RWA / Resident Welfare Associations", "city": "Delhi"},
    {"url": "https://es.scribd.com/document/543721592/RWA-List-South-Delhi", "category": "rwa", "category_name": "RWA / Resident Welfare Associations", "city": "Delhi"},
    # Government
    {"url": "https://delhishelterboard.in/main/wp-content/uploads/2019/06/Online-Directory-2016-Delhi-Govt.pdf", "category": "govt", "category_name": "Government & PSU Employees", "city": "Delhi"},
    {"url": "https://www.scribd.com/document/256455313/Delhi-ncr-Govt-Directory", "category": "govt", "category_name": "Government & PSU Employees", "city": "Delhi"},
    {"url": "https://www.scribd.com/document/832910190/Govt-Teachers-Delhi-Jeetu", "category": "schools", "category_name": "Schools & Teachers", "city": "Delhi"},
    {"url": "https://documents.doptcirculars.nic.in/D2/D02adm/RWA.pdf", "category": "govt", "category_name": "Government & PSU Employees", "city": "Delhi"},
    # Clubs
    {"url": "https://www.scribd.com/document/452887344/clubs-list-delhi", "category": "clubs", "category_name": "Club & Social Organizations", "city": "Delhi"},
    {"url": "https://ddca.in/assets/backend/uploade/ddca-document/2021/08/agmelections/List_of_Members.pdf", "category": "clubs", "category_name": "Club & Social Organizations", "city": "Delhi"},
    {"url": "https://e-clubhouse.org/sites/newdelhieast/page-7.php", "category": "clubs", "category_name": "Club & Social Organizations", "city": "Delhi"},
    {"url": "https://delhigymkhana.org.in/wp-content/uploads/2025/01/delhigymkahapdf.pdf", "category": "clubs", "category_name": "Club & Social Organizations", "city": "Delhi"},
    {"url": "https://www.scribd.com/document/377869455/Gymkhana-Delhi-Governing-Member-List", "category": "clubs", "category_name": "Club & Social Organizations", "city": "Delhi"},
    {"url": "https://rotaryindia.org/Documents/ebulletin/Group683/directory_with_data22072023032046PM.pdf", "category": "clubs", "category_name": "Club & Social Organizations", "city": "Delhi"},
    # Professional
    {"url": "https://www.iiipicai.in/wp-content/uploads/2023/06/Professional-Members-Directory-as-on-14-06-2023.pdf", "category": "professional", "category_name": "Professional Directories", "city": "Delhi"},
    {"url": "https://www.pwa.in/wp-content/uploads/2023/04/Members-Directory-April-2023.pdf", "category": "professional", "category_name": "Professional Directories", "city": "Delhi"},
    {"url": "https://ldo.gov.in/WriteReadData/UserFiles/file/DGC%20memb%20Nov-17.pdf", "category": "clubs", "category_name": "Club & Social Organizations", "city": "Delhi"},
    # Senior Citizens
    {"url": "https://www.esicncrpensioners.com/pdf/memberlist.pdf", "category": "senior", "category_name": "Senior Citizens & Elder Care", "city": "Delhi"},
    # PSU
    {"url": "https://www.scribd.com/document/434001974/PSU-List", "category": "govt", "category_name": "Government & PSU Employees", "city": "Delhi"},
]

# ============ Comprehensive location & entity lists ============
DELHI_COLONIES = [
    "Saket", "Sheikh Sarai", "Vasant Kunj", "RK Puram", "Greater Kailash I", "Greater Kailash II",
    "Defence Colony", "Lajpat Nagar", "South Extension", "Green Park", "Hauz Khas", "Safdarjung Enclave",
    "Malviya Nagar", "Chittaranjan Park", "Kalkaji", "Nehru Place", "East of Kailash",
    "Panchsheel Park", "Panchsheel Enclave", "Gulmohar Park", "Sarvapriya Vihar", "SDA",
    "Vasant Vihar", "Munirka", "Mehrauli", "Chattarpur", "Sainik Farm",
    "Dwarka Sector 1", "Dwarka Sector 2", "Dwarka Sector 3", "Dwarka Sector 4", "Dwarka Sector 5",
    "Dwarka Sector 6", "Dwarka Sector 7", "Dwarka Sector 8", "Dwarka Sector 9", "Dwarka Sector 10",
    "Dwarka Sector 11", "Dwarka Sector 12", "Dwarka Sector 13", "Dwarka Sector 14",
    "Dwarka Sector 18", "Dwarka Sector 19", "Dwarka Sector 21", "Dwarka Sector 22", "Dwarka Sector 23",
    "Rohini Sector 1", "Rohini Sector 2", "Rohini Sector 3", "Rohini Sector 4", "Rohini Sector 5",
    "Rohini Sector 7", "Rohini Sector 8", "Rohini Sector 9", "Rohini Sector 11", "Rohini Sector 13",
    "Rohini Sector 14", "Rohini Sector 15", "Rohini Sector 16", "Rohini Sector 24", "Rohini Sector 25",
    "Pitampura", "Shalimar Bagh", "Paschim Vihar", "Janakpuri", "Vikaspuri", "Uttam Nagar",
    "Rajouri Garden", "Tilak Nagar", "Punjabi Bagh", "Model Town", "Civil Lines",
    "Preet Vihar", "Laxmi Nagar", "Mayur Vihar Phase 1", "Mayur Vihar Phase 2", "Mayur Vihar Phase 3",
    "IP Extension", "Patparganj", "Vasundhara Enclave", "Vivek Vihar",
    "Sarita Vihar", "Jasola", "Sukhdev Vihar", "Okhla", "Jamia Nagar",
    "Alaknanda", "Navjivan Vihar", "Anand Niketan", "Westend Colony",
    "Maharani Bagh", "Friends Colony", "Nizamuddin", "Jor Bagh", "Sundar Nagar",
    "Naraina", "Moti Bagh", "Nauroji Nagar", "Kidwai Nagar", "Netaji Nagar",
    "Lodhi Colony", "Pandara Road", "Jangpura", "Andrews Ganj",
    "Yamuna Vihar", "Nand Nagri", "Dilshad Garden", "Shahdara",
]

GURGAON_AREAS = [
    "Sector 1", "Sector 2", "Sector 3", "Sector 4", "Sector 5", "Sector 7", "Sector 9",
    "Sector 10", "Sector 14", "Sector 15", "Sector 17", "Sector 21", "Sector 22", "Sector 23",
    "Sector 28", "Sector 29", "Sector 31", "Sector 33", "Sector 38", "Sector 39",
    "Sector 40", "Sector 42", "Sector 43", "Sector 44", "Sector 45", "Sector 46",
    "Sector 47", "Sector 48", "Sector 49", "Sector 50", "Sector 51", "Sector 52", "Sector 53", "Sector 54",
    "Sector 55", "Sector 56", "Sector 57", "Sector 58", "Sector 62",
    "Sector 67", "Sector 69", "Sector 70", "Sector 72", "Sector 76", "Sector 77", "Sector 80",
    "Sector 81", "Sector 82", "Sector 83", "Sector 84", "Sector 85", "Sector 86",
    "DLF Phase 1", "DLF Phase 2", "DLF Phase 3", "DLF Phase 4", "DLF Phase 5",
    "Sushant Lok Phase 1", "Sushant Lok Phase 2", "Sushant Lok Phase 3",
    "South City 1", "South City 2", "Nirvana Country", "Palam Vihar",
    "Malibu Town", "Suncity", "Ardee City", "Vatika City", "Emaar Palm Hills",
    "Golf Course Road", "MG Road", "Sohna Road",
]

NOIDA_AREAS = [
    "Sector 1", "Sector 2", "Sector 3", "Sector 4", "Sector 5", "Sector 6", "Sector 7",
    "Sector 8", "Sector 9", "Sector 10", "Sector 11", "Sector 12", "Sector 14", "Sector 15",
    "Sector 16", "Sector 17", "Sector 18", "Sector 19", "Sector 20", "Sector 21", "Sector 22", "Sector 23",
    "Sector 25", "Sector 27", "Sector 29", "Sector 30", "Sector 31", "Sector 32", "Sector 33",
    "Sector 34", "Sector 35", "Sector 36", "Sector 37", "Sector 38", "Sector 39",
    "Sector 40", "Sector 41", "Sector 44", "Sector 45", "Sector 46", "Sector 47", "Sector 48",
    "Sector 49", "Sector 50", "Sector 51", "Sector 52", "Sector 53", "Sector 54", "Sector 55", "Sector 56",
    "Sector 57", "Sector 58", "Sector 59", "Sector 60", "Sector 61", "Sector 62", "Sector 63",
    "Sector 70", "Sector 71", "Sector 72", "Sector 73", "Sector 74", "Sector 75", "Sector 76",
    "Sector 77", "Sector 78", "Sector 79", "Sector 82",
    "Sector 92", "Sector 93", "Sector 100", "Sector 104", "Sector 107",
    "Sector 110", "Sector 115", "Sector 117", "Sector 118", "Sector 119",
    "Sector 120", "Sector 121", "Sector 122", "Sector 128", "Sector 129",
    "Sector 130", "Sector 131", "Sector 132", "Sector 133", "Sector 134", "Sector 135",
    "Sector 137", "Sector 140", "Sector 142", "Sector 143", "Sector 144", "Sector 150",
]

CLUBS_DELHI_NCR = [
    "Delhi Gymkhana Club", "Chelmsford Club", "India International Centre",
    "India Habitat Centre", "Delhi Golf Club", "Roshanara Club",
    "National Sports Club of India", "Panchsheel Club", "Jahanpanah Club",
    "Siri Fort Sports Complex", "DDA Sports Complex", "ITC Maurya Club",
    "Cosmopolitan Club Delhi", "Press Club of India", "PHD Chamber",
    "Lions Club Delhi", "Rotary Club Delhi", "Rotary Club Gurgaon",
    "Rotary Club Noida", "Lions Club Gurgaon", "Lions Club Noida",
    "DLF Club", "DLF Golf and Country Club", "Karma Lakelands Club",
    "Heritage Village Club", "Aravali Golf Club", "Classic Golf Resort",
    "Golden Greens Golf Club", "Noida Golf Course",
    "IIT Delhi Alumni Association", "IIM Alumni Association Delhi",
    "All India Management Association", "PHD Chamber of Commerce",
    "FICCI", "CII", "ASSOCHAM",
]

PSU_COMPANIES = [
    "NTPC", "ONGC", "BHEL", "GAIL", "Indian Oil", "BPCL", "HPCL",
    "Power Grid Corporation", "Coal India", "SAIL", "NHPC",
    "NMDC", "NALCO", "CONCOR", "NBCC", "IRCON", "RITES",
    "Engineers India Limited", "THDC India", "SJVN", "NLC India",
    "Power Finance Corporation", "REC Limited", "IREDA",
    "LIC", "GIC", "SBI", "Bank of Baroda", "Punjab National Bank",
    "Canara Bank", "Union Bank", "Indian Bank", "Central Bank",
    "Air India", "Indian Railways", "DMRC", "AAI",
    "MTNL", "BSNL", "ITI Limited",
    "HAL", "BEL", "BEML", "BDL", "MDL", "Garden Reach Shipbuilders",
    "DRDO", "ISRO",
    "Food Corporation of India", "Central Warehousing Corporation",
    "HUDCO", "NBCC", "CPWD",
]

GOVT_DEPARTMENTS = [
    "DDA", "MCD", "NDMC", "Delhi Jal Board", "Delhi Police",
    "Delhi Development Authority", "DERC", "Delhi Metro",
    "Delhi Transport Corporation", "Delhi Fire Service",
    "Income Tax Department Delhi", "Customs Delhi",
    "CGHS Delhi", "ESIC Delhi", "EPF Delhi",
    "Noida Authority", "Greater Noida Authority", "YEIDA",
    "HUDA Gurgaon", "MCG Gurgaon", "Haryana Police Gurgaon",
    "District Court Delhi", "High Court Delhi",
]


def build_search_queries():
    """Generate massive search query list covering all categories, localities, entities."""
    queries = []
    seen = set()

    def add(q, cat_id, cat_name, city):
        if q not in seen:
            seen.add(q)
            queries.append({"query": q, "category": cat_id, "category_name": cat_name, "city": city, "filetype": "web"})

    # 1. RWA queries — every colony/sector
    for colony in DELHI_COLONIES:
        add(f"RWA {colony} Delhi members phone directory", "rwa", "RWA / Resident Welfare Associations", "Delhi")
        add(f"resident welfare association {colony} contact list", "rwa", "RWA / Resident Welfare Associations", "Delhi")

    for area in GURGAON_AREAS:
        add(f"RWA {area} Gurgaon members phone directory", "rwa", "RWA / Resident Welfare Associations", "Gurgaon")
        add(f"resident welfare association {area} Gurgaon contact", "rwa", "RWA / Resident Welfare Associations", "Gurgaon")

    for area in NOIDA_AREAS:
        add(f"RWA {area} Noida members phone directory", "rwa", "RWA / Resident Welfare Associations", "Noida")
        add(f"resident welfare association {area} Noida contact", "rwa", "RWA / Resident Welfare Associations", "Noida")

    # General RWA
    for city in CITIES:
        add(f"RWA directory {city} phone numbers list", "rwa", "RWA / Resident Welfare Associations", city)
        add(f"RWA office bearers {city} contact mobile", "rwa", "RWA / Resident Welfare Associations", city)
        add(f"RWA president secretary {city} phone list", "rwa", "RWA / Resident Welfare Associations", city)
        add(f"residential welfare association {city} members directory pdf", "rwa", "RWA / Resident Welfare Associations", city)
        add(f"housing society directory {city} phone numbers", "rwa", "RWA / Resident Welfare Associations", city)
        add(f"CGHS society {city} members list contact", "rwa", "RWA / Resident Welfare Associations", city)
        add(f"cooperative housing society {city} residents directory", "rwa", "RWA / Resident Welfare Associations", city)

    # 2. Club queries — every named club
    for club in CLUBS_DELHI_NCR:
        add(f"{club} members directory phone list", "clubs", "Club & Social Organizations", "Delhi")
        add(f"{club} member contact numbers", "clubs", "Club & Social Organizations", "Delhi")

    for city in CITIES:
        add(f"club members directory {city} phone list", "clubs", "Club & Social Organizations", city)
        add(f"sports club members list {city} contact", "clubs", "Club & Social Organizations", city)
        add(f"social club directory {city} members phone", "clubs", "Club & Social Organizations", city)

    # 3. PSU employee queries
    for psu in PSU_COMPANIES:
        add(f"{psu} employees directory phone list Delhi", "govt", "Government & PSU Employees", "Delhi")
        add(f"{psu} staff contact numbers list", "govt", "Government & PSU Employees", "Delhi")
        add(f"{psu} officers directory mobile number", "govt", "Government & PSU Employees", "Delhi")

    # 4. Government department queries
    for dept in GOVT_DEPARTMENTS:
        add(f"{dept} officers directory phone list", "govt", "Government & PSU Employees", "Delhi")
        add(f"{dept} employees contact numbers list", "govt", "Government & PSU Employees", "Delhi")
        add(f"{dept} staff directory mobile", "govt", "Government & PSU Employees", "Delhi")

    # 5. Schools & teachers
    for city in CITIES:
        add(f"school teachers directory {city} phone numbers", "schools", "Schools & Teachers", city)
        add(f"government school teachers list {city} contact", "schools", "Schools & Teachers", city)
        add(f"private school principals {city} phone directory", "schools", "Schools & Teachers", city)
        add(f"education department {city} teachers list mobile", "schools", "Schools & Teachers", city)
        add(f"college faculty directory {city} phone", "schools", "Schools & Teachers", city)
        add(f"university professors list {city} contact numbers", "schools", "Schools & Teachers", city)

    # 6. Professional directories
    for city in CITIES:
        add(f"chartered accountants directory {city} phone", "professional", "Professional Directories", city)
        add(f"advocates lawyers list {city} contact numbers", "professional", "Professional Directories", city)
        add(f"doctors directory {city} phone numbers list", "professional", "Professional Directories", city)
        add(f"architects directory {city} members phone", "professional", "Professional Directories", city)
        add(f"bar council members {city} contact list", "professional", "Professional Directories", city)
        add(f"engineers association {city} members directory", "professional", "Professional Directories", city)
        add(f"dental association {city} members phone list", "professional", "Professional Directories", city)

    # 7. Senior citizens
    for city in CITIES:
        add(f"senior citizens association {city} members phone", "senior", "Senior Citizens & Elder Care", city)
        add(f"retired employees association {city} directory", "senior", "Senior Citizens & Elder Care", city)
        add(f"pensioners association {city} contact list", "senior", "Senior Citizens & Elder Care", city)
        add(f"ex servicemen association {city} members phone", "senior", "Senior Citizens & Elder Care", city)
        add(f"retired government officers {city} directory", "senior", "Senior Citizens & Elder Care", city)
        add(f"old age welfare association {city} phone", "senior", "Senior Citizens & Elder Care", city)

    # 8. Religious & community
    for city in CITIES:
        add(f"gurdwara committee {city} members phone list", "religious", "Religious & Community Organizations", city)
        add(f"temple trust {city} committee members contact", "religious", "Religious & Community Organizations", city)
        add(f"masjid committee {city} members directory", "religious", "Religious & Community Organizations", city)
        add(f"church parish {city} members phone directory", "religious", "Religious & Community Organizations", city)

    # 9. Business & trade
    for city in CITIES:
        add(f"traders association {city} members phone list", "business", "Business & Trade Associations", city)
        add(f"market association {city} shopkeepers directory", "business", "Business & Trade Associations", city)
        add(f"chamber of commerce {city} members contact", "business", "Business & Trade Associations", city)
        add(f"industrial association {city} directory phone", "business", "Business & Trade Associations", city)

    return queries


def extract_phones_from_text(text):
    """Extract unique Indian phone numbers from text."""
    matches = PHONE_REGEX.findall(text)
    phones = set()
    for m in matches:
        phone = m.strip()
        if len(phone) == 10:
            phones.add(phone)
    return list(phones)


def extract_names_near_phones(text, phones):
    """Try to extract names near phone numbers."""
    contacts = []
    lines = text.split('\n')
    for i, line in enumerate(lines):
        for phone in phones:
            if phone in line:
                # Try to find a name in the same line or nearby lines
                name = None
                # Check current line for name-like pattern
                name_match = re.search(r'(?:Mr\.?|Mrs\.?|Ms\.?|Dr\.?|Sh\.?|Smt\.?|Shri\.?)?\s*([A-Z][a-z]+(?:\s+[A-Z][a-z]+){1,3})', line)
                if name_match:
                    name = name_match.group(0).strip()
                elif i > 0:
                    prev_match = re.search(r'(?:Mr\.?|Mrs\.?|Ms\.?|Dr\.?|Sh\.?|Smt\.?|Shri\.?)?\s*([A-Z][a-z]+(?:\s+[A-Z][a-z]+){1,3})', lines[i-1])
                    if prev_match:
                        name = prev_match.group(0).strip()
                
                # Extract email if present
                email_match = re.search(r'[\w.+-]+@[\w-]+\.[\w.-]+', line)
                email = email_match.group(0) if email_match else None
                
                # Extract address hints
                address = None
                for addr_line in [line] + (lines[max(0,i-2):i] if i > 0 else []):
                    if any(kw in addr_line.lower() for kw in ['sector', 'block', 'colony', 'nagar', 'road', 'street', 'marg', 'vihar', 'enclave', 'garden', 'park', 'phase']):
                        address = addr_line.strip()[:200]
                        break

                contacts.append({
                    "phone": phone,
                    "name": name,
                    "email": email,
                    "address": address
                })
                break
    return contacts


def extract_from_pdf(filepath):
    """Extract contacts from a PDF file."""
    import pdfplumber
    contacts = []
    try:
        with pdfplumber.open(filepath) as pdf:
            full_text = ""
            for page in pdf.pages[:50]:  # Limit to 50 pages
                text = page.extract_text() or ""
                full_text += text + "\n"
                
                # Also try tables
                tables = page.extract_tables()
                for table in tables:
                    for row in table:
                        if row:
                            row_text = " ".join([str(cell) for cell in row if cell])
                            full_text += row_text + "\n"
            
            phones = extract_phones_from_text(full_text)
            if phones:
                contacts = extract_names_near_phones(full_text, phones)
                # For phones without extracted context, add them bare
                found_phones = {c["phone"] for c in contacts}
                for p in phones:
                    if p not in found_phones:
                        contacts.append({"phone": p, "name": None, "email": None, "address": None})
    except Exception as e:
        logger.error(f"PDF extraction error for {filepath}: {e}")
    return contacts


def extract_from_excel(filepath):
    """Extract contacts from Excel/CSV."""
    import pandas as pd
    contacts = []
    try:
        if str(filepath).endswith('.csv'):
            dfs = [pd.read_csv(filepath, encoding='utf-8', on_bad_lines='skip')]
        else:
            xls = pd.ExcelFile(filepath)
            dfs = [pd.read_excel(xls, sheet_name=s) for s in xls.sheet_names[:5]]
        
        for df in dfs:
            df = df.fillna('')
            full_text = df.to_string()
            phones = extract_phones_from_text(full_text)
            
            # Try column-based extraction
            phone_cols = [c for c in df.columns if any(kw in str(c).lower() for kw in ['phone', 'mobile', 'contact', 'cell', 'mob', 'tel'])]
            name_cols = [c for c in df.columns if any(kw in str(c).lower() for kw in ['name', 'person', 'member', 'resident'])]
            email_cols = [c for c in df.columns if any(kw in str(c).lower() for kw in ['email', 'mail', 'e-mail'])]
            addr_cols = [c for c in df.columns if any(kw in str(c).lower() for kw in ['address', 'addr', 'location', 'sector', 'block', 'flat'])]
            
            if phone_cols:
                for _, row in df.iterrows():
                    for pc in phone_cols:
                        val = str(row[pc]).strip()
                        row_phones = extract_phones_from_text(val)
                        for p in row_phones:
                            name = str(row[name_cols[0]]).strip() if name_cols else None
                            email_val = str(row[email_cols[0]]).strip() if email_cols else None
                            addr = str(row[addr_cols[0]]).strip() if addr_cols else None
                            if name and name.lower() in ['nan', '', 'none']:
                                name = None
                            if email_val and email_val.lower() in ['nan', '', 'none']:
                                email_val = None
                            if addr and addr.lower() in ['nan', '', 'none']:
                                addr = None
                            contacts.append({"phone": p, "name": name, "email": email_val, "address": addr})
            else:
                # Fallback: regex scan all text
                for p in phones:
                    contacts.append({"phone": p, "name": None, "email": None, "address": None})
    except Exception as e:
        logger.error(f"Excel extraction error for {filepath}: {e}")
    return contacts


def run_crawl_worker(job_id):
    """Main crawl worker running in background thread."""
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    loop.run_until_complete(_async_crawl_worker(job_id))
    loop.close()


async def _async_crawl_worker(job_id):
    """Async crawl: search → download → extract → store. Uses DuckDuckGo to avoid Google rate limits."""
    import requests
    from duckduckgo_search import DDGS

    sync_client = MongoClient(mongo_url)
    sync_db = sync_client[db_name]

    all_queries = build_search_queries()
    
    # Resume: find which queries were already done
    done_queries = set()
    for doc in sync_db.hc_crawl_queries.find({"status": "done"}, {"query": 1}):
        done_queries.add(doc["query"])
    
    # Also track already-processed URLs
    processed_urls = set()
    for doc in sync_db.hc_sources.find({}, {"url": 1}):
        processed_urls.add(doc["url"])
    
    remaining = [q for q in all_queries if q["query"] not in done_queries]
    total_queries = len(all_queries)
    done_count = len(done_queries)

    active_crawl[job_id] = {
        "status": "running",
        "total_queries": total_queries,
        "queries_done": done_count,
        "queries_remaining": len(remaining),
        "files_found": 0,
        "files_downloaded": 0,
        "contacts_extracted": 0,
        "current_query": "",
        "started_at": datetime.now(timezone.utc).isoformat()
    }

    # Get running totals from DB
    total_contacts_db = sync_db.hc_leads.count_documents({})
    total_files_db = sync_db.hc_sources.count_documents({})
    active_crawl[job_id]["contacts_extracted"] = total_contacts_db
    active_crawl[job_id]["files_downloaded"] = total_files_db

    logger.info(f"HearClear crawl {job_id}: {len(remaining)} queries remaining of {total_queries}")

    # ========== PHASE 1: Process seed URLs first ==========
    for seed in SEED_URLS:
        if active_crawl.get(job_id, {}).get("status") != "running":
            break
        url = seed["url"]
        if url in processed_urls:
            continue
        processed_urls.add(url)
        active_crawl[job_id]["current_query"] = f"[SEED] {url[:60]}..."

        try:
            is_file = any(url.lower().endswith(f'.{ft}') for ft in FILE_TYPES)

            resp = requests.get(url, timeout=30, headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
            }, allow_redirects=True)
            if resp.status_code != 200 or len(resp.content) < 200:
                continue

            contacts = []
            filename = None

            if is_file:
                ext = "pdf"
                for ft in FILE_TYPES:
                    if url.lower().endswith(f'.{ft}'):
                        ext = ft
                        break
                filename = f"seed_{uuid.uuid4().hex[:12]}.{ext}"
                filepath = DOWNLOADS_DIR / filename
                with open(filepath, "wb") as f:
                    f.write(resp.content)
                if ext == "pdf":
                    contacts = extract_from_pdf(str(filepath))
                elif ext in ["xls", "xlsx", "csv"]:
                    contacts = extract_from_excel(str(filepath))
            else:
                # Web page
                page_text = resp.text
                phones = extract_phones_from_text(page_text)
                if phones:
                    import re as re_mod
                    clean_text = re_mod.sub(r'<[^>]+>', ' ', page_text)
                    contacts = extract_names_near_phones(clean_text, phones)
                    found_phones = {c["phone"] for c in contacts}
                    for p in phones:
                        if p not in found_phones:
                            contacts.append({"phone": p, "name": None, "email": None, "address": None})

            if contacts:
                source_doc = {
                    "source_id": str(uuid.uuid4()),
                    "url": url,
                    "filename": filename,
                    "file_type": "pdf" if is_file else "web",
                    "file_size": len(resp.content),
                    "category": seed["category"],
                    "category_name": seed["category_name"],
                    "city": seed["city"],
                    "query": "SEED",
                    "contacts_found": len(contacts),
                    "downloaded_at": datetime.now(timezone.utc).isoformat()
                }
                sync_db.hc_sources.insert_one(source_doc)
                active_crawl[job_id]["files_downloaded"] += 1

                new_contacts = 0
                for contact in contacts:
                    phone = contact["phone"]
                    if not sync_db.hc_leads.find_one({"phone": phone}):
                        sync_db.hc_leads.insert_one({
                            "phone": phone,
                            "name": contact.get("name"),
                            "email": contact.get("email"),
                            "address": contact.get("address"),
                            "source_url": url,
                            "source_file": filename,
                            "category": seed["category"],
                            "category_name": seed["category_name"],
                            "city": seed["city"],
                            "added_at": datetime.now(timezone.utc).isoformat()
                        })
                        new_contacts += 1

                active_crawl[job_id]["contacts_extracted"] = sync_db.hc_leads.count_documents({})
                logger.info(f"SEED {url[:60]}: {new_contacts} new contacts from {len(contacts)} found")

        except Exception as e:
            logger.debug(f"Seed URL error {url[:60]}: {e}")

    # ========== PHASE 2: DuckDuckGo search queries ==========
    ddgs = DDGS()

    for qi, qdata in enumerate(remaining):
        if active_crawl.get(job_id, {}).get("status") != "running":
            break

        query = qdata["query"]
        active_crawl[job_id]["current_query"] = query
        active_crawl[job_id]["queries_remaining"] = len(remaining) - qi

        try:
            # Search via DuckDuckGo
            urls_found = []
            web_pages = []
            try:
                results = ddgs.text(query, max_results=15, region="in-en")
                for r in results:
                    href = r.get("href", "")
                    lower_href = href.lower()
                    # Downloadable files
                    if any(lower_href.endswith(f'.{ft}') for ft in FILE_TYPES):
                        urls_found.append(href)
                    elif any(f'.{ft}?' in lower_href or f'.{ft}#' in lower_href for ft in FILE_TYPES):
                        urls_found.append(href.split('?')[0].split('#')[0])
                    else:
                        # Regular web page — scrape for phone numbers
                        web_pages.append(href)
                
                time.sleep(3)
            except Exception as e:
                logger.warning(f"DDG search error for '{query}': {e}")
                time.sleep(10)

            active_crawl[job_id]["files_found"] += len(urls_found)

            for url in urls_found:
                if active_crawl.get(job_id, {}).get("status") != "running":
                    break

                # Skip if already processed
                if url in processed_urls:
                    continue
                processed_urls.add(url)

                try:
                    # Download file
                    resp = requests.get(url, timeout=30, headers={
                        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
                    }, allow_redirects=True)
                    if resp.status_code != 200 or len(resp.content) < 500:
                        continue

                    # Determine file type from URL or content-type
                    ext = "pdf"
                    content_type = resp.headers.get("content-type", "").lower()
                    for ft in FILE_TYPES:
                        if url.lower().endswith(f'.{ft}'):
                            ext = ft
                            break
                    if "spreadsheet" in content_type or "excel" in content_type:
                        ext = "xlsx"
                    elif "csv" in content_type:
                        ext = "csv"
                    
                    filename = f"{uuid.uuid4().hex[:12]}.{ext}"
                    filepath = DOWNLOADS_DIR / filename
                    with open(filepath, "wb") as f:
                        f.write(resp.content)

                    # Extract contacts
                    contacts = []
                    if ext == "pdf":
                        contacts = extract_from_pdf(str(filepath))
                    elif ext in ["xls", "xlsx", "csv"]:
                        contacts = extract_from_excel(str(filepath))

                    # Store source record
                    source_doc = {
                        "source_id": str(uuid.uuid4()),
                        "url": url,
                        "filename": filename,
                        "file_type": ext,
                        "file_size": len(resp.content),
                        "category": qdata["category"],
                        "category_name": qdata["category_name"],
                        "city": qdata["city"],
                        "query": query,
                        "contacts_found": len(contacts),
                        "downloaded_at": datetime.now(timezone.utc).isoformat()
                    }
                    sync_db.hc_sources.insert_one(source_doc)
                    active_crawl[job_id]["files_downloaded"] += 1

                    # Store contacts (deduplicate on phone)
                    new_contacts = 0
                    for contact in contacts:
                        phone = contact["phone"]
                        existing = sync_db.hc_leads.find_one({"phone": phone})
                        if not existing:
                            lead_doc = {
                                "phone": phone,
                                "name": contact.get("name"),
                                "email": contact.get("email"),
                                "address": contact.get("address"),
                                "source_url": url,
                                "source_file": filename,
                                "category": qdata["category"],
                                "category_name": qdata["category_name"],
                                "city": qdata["city"],
                                "added_at": datetime.now(timezone.utc).isoformat()
                            }
                            sync_db.hc_leads.insert_one(lead_doc)
                            new_contacts += 1
                        else:
                            # Update with additional info if we have better data
                            updates = {}
                            if contact.get("name") and not existing.get("name"):
                                updates["name"] = contact["name"]
                            if contact.get("email") and not existing.get("email"):
                                updates["email"] = contact["email"]
                            if contact.get("address") and not existing.get("address"):
                                updates["address"] = contact["address"]
                            if updates:
                                sync_db.hc_leads.update_one({"phone": phone}, {"$set": updates})

                    active_crawl[job_id]["contacts_extracted"] = sync_db.hc_leads.count_documents({})
                    
                    if new_contacts > 0:
                        logger.info(f"Source {url}: {new_contacts} new contacts from {len(contacts)} total")

                except Exception as e:
                    logger.debug(f"Download/extract error for {url}: {e}")
                    continue

            # Also scrape web pages for phone numbers directly
            for page_url in web_pages[:5]:  # Limit to 5 pages per query
                if active_crawl.get(job_id, {}).get("status") != "running":
                    break
                if page_url in processed_urls:
                    continue
                processed_urls.add(page_url)

                try:
                    resp = requests.get(page_url, timeout=15, headers={
                        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
                    }, allow_redirects=True)
                    if resp.status_code != 200:
                        continue
                    
                    page_text = resp.text
                    # Extract phone numbers from HTML text
                    phones = extract_phones_from_text(page_text)
                    if not phones:
                        continue

                    # Clean HTML to get text for name extraction
                    import re as re_mod
                    clean_text = re_mod.sub(r'<[^>]+>', ' ', page_text)
                    contacts = extract_names_near_phones(clean_text, phones)
                    found_phones = {c["phone"] for c in contacts}
                    for p in phones:
                        if p not in found_phones:
                            contacts.append({"phone": p, "name": None, "email": None, "address": None})

                    if contacts:
                        # Store source
                        source_doc = {
                            "source_id": str(uuid.uuid4()),
                            "url": page_url,
                            "filename": None,
                            "file_type": "web",
                            "file_size": len(resp.content),
                            "category": qdata["category"],
                            "category_name": qdata["category_name"],
                            "city": qdata["city"],
                            "query": query,
                            "contacts_found": len(contacts),
                            "downloaded_at": datetime.now(timezone.utc).isoformat()
                        }
                        sync_db.hc_sources.insert_one(source_doc)
                        active_crawl[job_id]["files_downloaded"] += 1

                        new_contacts = 0
                        for contact in contacts:
                            phone = contact["phone"]
                            existing = sync_db.hc_leads.find_one({"phone": phone})
                            if not existing:
                                lead_doc = {
                                    "phone": phone,
                                    "name": contact.get("name"),
                                    "email": contact.get("email"),
                                    "address": contact.get("address"),
                                    "source_url": page_url,
                                    "source_file": None,
                                    "category": qdata["category"],
                                    "category_name": qdata["category_name"],
                                    "city": qdata["city"],
                                    "added_at": datetime.now(timezone.utc).isoformat()
                                }
                                sync_db.hc_leads.insert_one(lead_doc)
                                new_contacts += 1

                        active_crawl[job_id]["contacts_extracted"] = sync_db.hc_leads.count_documents({})
                        if new_contacts > 0:
                            logger.info(f"Web page {page_url}: {new_contacts} new contacts from {len(contacts)} found")

                except Exception as e:
                    logger.debug(f"Web scrape error for {page_url}: {e}")
                    continue

        except Exception as e:
            logger.error(f"Query processing error: {e}")

        # Mark query as done
        sync_db.hc_crawl_queries.update_one(
            {"query": query},
            {"$set": {"query": query, "status": "done", "completed_at": datetime.now(timezone.utc).isoformat()}},
            upsert=True
        )
        active_crawl[job_id]["queries_done"] += 1

    sync_client.close()

    if active_crawl.get(job_id, {}).get("status") == "running":
        active_crawl[job_id]["status"] = "completed"
    active_crawl[job_id]["completed_at"] = datetime.now(timezone.utc).isoformat()
    logger.info(f"HearClear crawl {job_id} done: {active_crawl[job_id]['contacts_extracted']} total contacts")


# ============== API Endpoints ==============
@router.post("/start")
async def start_crawl():
    """Start or resume the HearClear lead crawler."""
    for jid, job in active_crawl.items():
        if job["status"] == "running":
            return {"message": "Crawl already running", "job_id": jid, "status": job}

    job_id = str(uuid.uuid4())

    thread = threading.Thread(target=run_crawl_worker, args=(job_id,), daemon=True)
    thread.start()

    return {"job_id": job_id, "status": "started"}


@router.post("/stop")
async def stop_crawl():
    for jid, job in active_crawl.items():
        if job["status"] == "running":
            job["status"] = "stopped"
            return {"message": "Crawl stopped", "job_id": jid}
    return {"message": "No active crawl"}


@router.get("/status")
async def get_crawl_status():
    if active_crawl:
        latest = max(active_crawl.keys(), key=lambda k: active_crawl[k].get("started_at", ""))
        return active_crawl[latest]

    total_queries = len(build_search_queries())
    done_queries = await db.hc_crawl_queries.count_documents({"status": "done"})
    total_contacts = await db.hc_leads.count_documents({})
    total_sources = await db.hc_sources.count_documents({})

    return {
        "status": "idle" if done_queries == 0 else "resumable",
        "total_queries": total_queries,
        "queries_done": done_queries,
        "queries_remaining": total_queries - done_queries,
        "files_downloaded": total_sources,
        "contacts_extracted": total_contacts
    }


@router.get("/leads")
async def get_leads(
    category: str = Query(None),
    city: str = Query(None),
    search: str = Query(None),
    page: int = 1,
    per_page: int = 100
):
    """Get extracted leads with filters."""
    query = {}
    if category:
        query["category"] = category
    if city:
        query["city"] = {"$regex": city, "$options": "i"}
    if search:
        query["$or"] = [
            {"phone": {"$regex": search}},
            {"name": {"$regex": search, "$options": "i"}},
            {"email": {"$regex": search, "$options": "i"}}
        ]

    skip = (page - 1) * per_page
    leads = await db.hc_leads.find(query, {"_id": 0}).sort("added_at", -1).skip(skip).limit(per_page).to_list(per_page)
    total = await db.hc_leads.count_documents(query)

    return {"leads": leads, "total": total, "page": page, "per_page": per_page}


@router.get("/sources")
async def get_sources(limit: int = 50):
    """Get crawled source files."""
    sources = await db.hc_sources.find({}, {"_id": 0}).sort("downloaded_at", -1).to_list(limit)
    return {"sources": sources}


@router.get("/stats")
async def get_lead_stats():
    """Get aggregate stats by category and city."""
    pipeline_cat = [{"$group": {"_id": "$category_name", "count": {"$sum": 1}}}]
    pipeline_city = [{"$group": {"_id": "$city", "count": {"$sum": 1}}}]

    cat_stats = await db.hc_leads.aggregate(pipeline_cat).to_list(50)
    city_stats = await db.hc_leads.aggregate(pipeline_city).to_list(50)

    total = await db.hc_leads.count_documents({})
    with_name = await db.hc_leads.count_documents({"name": {"$ne": None}})
    with_email = await db.hc_leads.count_documents({"email": {"$ne": None}})

    return {
        "total_leads": total,
        "with_name": with_name,
        "with_email": with_email,
        "by_category": {s["_id"]: s["count"] for s in cat_stats if s["_id"]},
        "by_city": {s["_id"]: s["count"] for s in city_stats if s["_id"]}
    }


@router.get("/download")
async def download_leads(category: str = Query(None), city: str = Query(None)):
    """Download leads as Excel."""
    query = {}
    if category:
        query["category"] = category
    if city:
        query["city"] = {"$regex": city, "$options": "i"}

    leads = await db.hc_leads.find(query, {"_id": 0}).sort("phone", 1).to_list(500000)
    if not leads:
        raise HTTPException(status_code=404, detail="No leads found")

    wb = Workbook()
    ws = wb.active
    ws.title = "HearClear Leads"
    header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='0A1628', end_color='0A1628', fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    headers = ['#', 'Phone', 'Name', 'Email', 'Address', 'Category', 'City', 'Source']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for idx, lead in enumerate(leads, 2):
        values = [
            idx - 1,
            lead.get('phone', ''),
            lead.get('name', ''),
            lead.get('email', ''),
            lead.get('address', ''),
            lead.get('category_name', ''),
            lead.get('city', ''),
            lead.get('source_url', '')
        ]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=idx, column=col, value=v or '')
            cell.border = thin_border

    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 50
    ws.freeze_panes = 'A2'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"HearClear_Leads_Delhi_NCR_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/categories")
async def get_categories():
    return {"categories": [{"id": c["id"], "name": c["name"]} for c in CATEGORIES]}
