import os
import re
import uuid
import httpx
import logging
import io
import asyncio
import threading
from datetime import datetime, timezone
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional, List
from motor.motor_asyncio import AsyncIOMotorClient
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/directory", tags=["directory"])

# MongoDB
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]


def clean_pdf_text(text):
    """Clean spacing artifacts from PDF text extraction."""
    if not text:
        return ""
    # Common PDF encoding splits - rejoin known words
    # Remove extra internal spaces while preserving actual word boundaries
    text = re.sub(r'\\r\\n|\\r|\\n', ' ', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def parse_exhibitor_page(text):
    """Extract structured data from a single exhibitor page using position-based parsing."""
    result = {}

    field_patterns = [
        ('name', r'Name\s+of\s+the\s+E\s*xhibit\s*or\s*:'),
        ('address', r'Addr\s*ess\s*:'),
        ('contact_person', r'Cont\s*act\s*Person\s*:'),
        ('designation', r'Designa\s*tio\s*n?\s*:'),
        ('phone', r'Phone\s*:'),
        ('mobile', r'Mobile\s*:'),
        ('email', r'Email\s*:'),
        ('website', r'Websit\s*e\s*:'),
        ('profile', r'Profile\s*:'),
    ]

    # Find positions of all field labels
    field_positions = []
    for field_name, pattern in field_patterns:
        m = re.search(pattern, text, re.IGNORECASE)
        if m:
            field_positions.append((m.end(), field_name, m.start()))

    field_positions.sort(key=lambda x: x[0])

    for i, (label_end, field_name, label_start) in enumerate(field_positions):
        if i + 1 < len(field_positions):
            next_label_start = field_positions[i + 1][2]
            value = text[label_end:next_label_start].strip()
        else:
            value = text[label_end:].strip()

        value = clean_pdf_text(value)

        if field_name in ('email', 'website'):
            value = re.sub(r'\s+', '', value)

        if field_name == 'profile':
            value = value[:1000]

        result[field_name] = value

    return result


def extract_all_exhibitors(pdf_bytes):
    """Extract all exhibitor records from the PDF using pdfplumber for clean text."""
    exhibitors = []

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        # Skip first 2 pages (cover + partners)
        for i in range(2, len(pdf.pages)):
            try:
                text = pdf.pages[i].extract_text()
                if not text:
                    continue

                data = parse_exhibitor_page(text)
                if not data.get('name'):
                    continue

                data['page_number'] = i + 1
                data['id'] = str(uuid.uuid4())
                exhibitors.append(data)
            except Exception as e:
                logger.error(f"Error parsing page {i+1}: {e}")

    return exhibitors


# Track extraction jobs
extraction_jobs = {}


def run_extraction_sync(job_id, pdf_bytes):
    """Run extraction in a thread (pdfplumber is CPU-bound)."""
    from motor.motor_asyncio import AsyncIOMotorClient as MC
    import asyncio

    try:
        exhibitors = extract_all_exhibitors(pdf_bytes)
        if not exhibitors:
            extraction_jobs[job_id] = {"status": "failed", "error": "No exhibitor data found"}
            return

        # Use sync pymongo for thread
        from pymongo import MongoClient
        sync_client = MongoClient(os.environ['MONGO_URL'])
        sync_db = sync_client[os.environ['DB_NAME']]

        batch_id = str(uuid.uuid4())
        now = datetime.now(timezone.utc).isoformat()
        docs = []
        for ex in exhibitors:
            docs.append({"batch_id": batch_id, "extracted_at": now, **ex})

        sync_db.exhibitors.delete_many({})
        sync_db.exhibitors.insert_many(docs)
        sync_client.close()

        extraction_jobs[job_id] = {
            "status": "completed",
            "count": len(exhibitors),
            "batch_id": batch_id
        }
        logger.info(f"Extraction job {job_id} completed: {len(exhibitors)} companies")

    except Exception as e:
        logger.error(f"Extraction job {job_id} failed: {e}")
        extraction_jobs[job_id] = {"status": "failed", "error": str(e)}


# ============== API Endpoints ==============
class ExtractRequest(BaseModel):
    pdf_url: str


@router.post("/extract-from-url")
async def extract_from_url(request: ExtractRequest):
    """Download PDF and start extraction in background."""
    try:
        async with httpx.AsyncClient(timeout=60, follow_redirects=True) as http_client:
            resp = await http_client.get(request.pdf_url)
            if resp.status_code != 200:
                raise HTTPException(status_code=400, detail="Failed to download PDF")
            pdf_bytes = resp.content

        job_id = str(uuid.uuid4())
        extraction_jobs[job_id] = {"status": "processing"}

        # Run in background thread
        thread = threading.Thread(target=run_extraction_sync, args=(job_id, pdf_bytes))
        thread.start()

        return {"job_id": job_id, "status": "processing", "message": "Extraction started"}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Extraction start error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/extract-status/{job_id}")
async def get_extraction_status(job_id: str):
    """Poll extraction job status."""
    job = extraction_jobs.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    return job


@router.post("/extract-upload")
async def extract_from_upload(file: UploadFile = File(...)):
    """Upload PDF and start extraction in background."""
    if not file.filename.lower().endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Only PDF files are supported")

    try:
        pdf_bytes = await file.read()

        job_id = str(uuid.uuid4())
        extraction_jobs[job_id] = {"status": "processing"}

        thread = threading.Thread(target=run_extraction_sync, args=(job_id, pdf_bytes))
        thread.start()

        return {"job_id": job_id, "status": "processing", "message": "Extraction started"}

    except Exception as e:
        logger.error(f"Upload extraction error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/companies")
async def get_companies(search: str = Query(None), page: int = 1, per_page: int = 200):
    """Get all extracted companies with optional search."""
    query = {}
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"contact_person": {"$regex": search, "$options": "i"}},
            {"email": {"$regex": search, "$options": "i"}},
            {"address": {"$regex": search, "$options": "i"}},
            {"profile": {"$regex": search, "$options": "i"}}
        ]

    total = await db.exhibitors.count_documents(query)
    skip = (page - 1) * per_page

    companies = await db.exhibitors.find(
        query, {"_id": 0}
    ).sort("page_number", 1).skip(skip).limit(per_page).to_list(per_page)

    return {
        "companies": companies,
        "total": total,
        "page": page,
        "per_page": per_page
    }


@router.get("/download-excel")
async def download_excel(search: str = Query(None)):
    """Download all extracted companies as Excel file."""
    query = {}
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"contact_person": {"$regex": search, "$options": "i"}},
            {"email": {"$regex": search, "$options": "i"}}
        ]

    companies = await db.exhibitors.find(
        query, {"_id": 0}
    ).sort("page_number", 1).to_list(500)

    if not companies:
        raise HTTPException(status_code=404, detail="No companies found")

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "G20 DIA Summit Exhibitors"

    # Header styling
    header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = ['#', 'Company Name', 'Address', 'Contact Person', 'Designation',
               'Phone', 'Mobile', 'Email', 'Website', 'Profile']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows
    data_alignment = Alignment(vertical='top', wrap_text=True)
    for row_idx, company in enumerate(companies, 2):
        values = [
            row_idx - 1,
            company.get('name', ''),
            company.get('address', ''),
            company.get('contact_person', ''),
            company.get('designation', ''),
            company.get('phone', ''),
            company.get('mobile', ''),
            company.get('email', ''),
            company.get('website', ''),
            company.get('profile', '')
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = data_alignment
            cell.border = thin_border

    # Column widths
    col_widths = [5, 35, 45, 25, 25, 18, 18, 30, 30, 60]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else 'A'].width = width

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Auto-filter
    ws.auto_filter.ref = f"A1:J{len(companies) + 1}"

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"G20_DIA_Exhibitors_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/stats")
async def get_stats():
    """Get extraction statistics."""
    total = await db.exhibitors.count_documents({})
    if total == 0:
        return {"total": 0, "extracted": False}

    sample = await db.exhibitors.find_one({}, {"_id": 0, "batch_id": 1, "extracted_at": 1})
    return {
        "total": total,
        "extracted": True,
        "batch_id": sample.get("batch_id") if sample else None,
        "extracted_at": sample.get("extracted_at") if sample else None
    }
