import os
import uuid
import logging
import threading
import asyncio
import io
from datetime import datetime, timezone
from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional, List
from motor.motor_asyncio import AsyncIOMotorClient
from pymongo import MongoClient
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/checker", tags=["checker"])

mongo_url = os.environ['MONGO_URL']
db_name = os.environ['DB_NAME']
client = AsyncIOMotorClient(mongo_url)
db = client[db_name]

EMAIL_RANGES = [
    {"prefix": "veenu", "start": 1, "end": 9999, "pad": 3, "domain": "gmail.com"},
    {"prefix": "vinty", "start": 300, "end": 1000, "pad": 0, "domain": "gmail.com"},
    {"prefix": "crazy", "start": 300, "end": 1000, "pad": 0, "domain": "gmail.com"},
    {"prefix": "strike", "start": 100, "end": 700, "pad": 0, "domain": "gmail.com"},
    {"prefix": "treaty", "start": 1, "end": 99, "pad": 4, "domain": "gmail.com"},
    {"prefix": "treaty", "start": 100, "end": 199, "pad": 4, "domain": "gmail.com"},
    {"prefix": "treaty", "start": 200, "end": 299, "pad": 4, "domain": "gmail.com"},
    {"prefix": "treaty", "start": 500, "end": 599, "pad": 4, "domain": "gmail.com"},
    {"prefix": "treaty", "start": 900, "end": 999, "pad": 4, "domain": "gmail.com"},
    {"prefix": "treaty", "start": 1000, "end": 1099, "pad": 4, "domain": "gmail.com"},
    {"prefix": "vineet", "start": 100, "end": 999, "pad": 0, "domain": "gmail.com"},
    {"prefix": "vineet", "start": 1000, "end": 1099, "pad": 0, "domain": "gmail.com"},
    {"prefix": "vineet", "start": 2000, "end": 2099, "pad": 0, "domain": "gmail.com"},
    {"prefix": "vineet", "start": 3000, "end": 3099, "pad": 0, "domain": "gmail.com"},
    {"prefix": "vineet", "start": 5000, "end": 5099, "pad": 0, "domain": "gmail.com"},
    {"prefix": "vineet", "start": 9000, "end": 9099, "pad": 0, "domain": "gmail.com"},
    {"prefix": "vngnara", "start": 500, "end": 1000, "pad": 0, "domain": "gmail.com"},
    {"prefix": "vininara", "start": 300, "end": 600, "pad": 0, "domain": "gmail.com"},
    {"prefix": "super", "start": 300, "end": 1100, "pad": 0, "domain": "gmail.com"},
    {"prefix": "vinsum", "start": 300, "end": 1099, "pad": 0, "domain": "gmail.com"},
]

# Skip these veenu ranges — no valid accounts expected
VEENU_SKIP_RANGES = [
    (1150, 1999),
    (2100, 2950),
    (3100, 3950),
    (4100, 4950),
    (5100, 5950),
    (6100, 6950),
    (7100, 7950),
    (8100, 8950),
    (9100, 9950),
]

PASSWORD = "c304i109"
active_jobs = {}
active_credits_jobs = {}


async def _scrape_credits(page, email, password):
    """Login and scrape credits from DDC by intercepting the lobby/game API. Returns credits string or None."""
    try:
        credits_result = {"value": None}
        
        async def capture_credits(response):
            url = response.url
            if 'lobby/game' in url:
                try:
                    text = await response.text()
                    import re as _re
                    cash_match = _re.search(r'"cash"\s*:\s*(\d+)', text)
                    psc_match = _re.search(r'"psc"\s*:\s*(\d+)', text)
                    if cash_match:
                        credits_result["value"] = cash_match.group(1)
                    elif psc_match:
                        credits_result["value"] = psc_match.group(1)
                except Exception:
                    pass

        page.on('response', capture_credits)
        
        await page.goto('https://www.doubledowncasino.com', wait_until='domcontentloaded', timeout=20000)
        await page.wait_for_timeout(2500)

        play_btn = await page.query_selector('img[src*="playnow"]')
        if play_btn:
            await play_btn.click(force=True)
            await page.wait_for_timeout(2000)

        try:
            await page.click('img[src*="email_connect"]', force=True, timeout=5000)
            await page.wait_for_timeout(800)
        except Exception:
            page.remove_listener('response', capture_credits)
            return None

        try:
            await page.click('img[src*="email_dailog_login"]', force=True, timeout=3000)
            await page.wait_for_timeout(400)
        except Exception:
            pass

        try:
            await page.fill('#emailID', email, timeout=3000)
            await page.fill('#pw', password, timeout=3000)
        except Exception:
            page.remove_listener('response', capture_credits)
            return None

        try:
            await page.click('img[src*="green_login"]', force=True, timeout=3000)
        except Exception:
            page.remove_listener('response', capture_credits)
            return None

        # Wait for the lobby/game API call that has the balance
        await page.wait_for_timeout(12000)
        
        page.remove_listener('response', capture_credits)
        return credits_result["value"]

    except Exception as e:
        logger.debug(f"Credits scrape error for {email}: {e}")
        return None


def run_credits_worker(job_id):
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    loop.run_until_complete(_async_credits_worker(job_id))
    loop.close()


async def _async_credits_worker(job_id):
    """Scrape credits for all successful accounts that don't have credits yet."""
    os.environ['PLAYWRIGHT_BROWSERS_PATH'] = '/pw-browsers'
    from playwright.async_api import async_playwright

    sync_client = MongoClient(mongo_url)
    sync_db = sync_client[db_name]

    # Get successful accounts without real credits (skip LOGIN_OK_CREDITS_UNKNOWN too)
    accounts = list(sync_db.login_results.find(
        {"status": "success", "$or": [
            {"credits": {"$exists": False}},
            {"credits": None},
            {"credits": "LOGIN_OK_CREDITS_UNKNOWN"}
        ]},
        {"email": 1}
    ))

    total = len(accounts)
    active_credits_jobs[job_id] = {
        "status": "running",
        "total": total,
        "processed": 0,
        "credits_found": 0,
        "current_email": "",
        "started_at": datetime.now(timezone.utc).isoformat()
    }

    logger.info(f"Credits scan {job_id}: {total} accounts to process")

    if total == 0:
        active_credits_jobs[job_id]["status"] = "completed"
        sync_client.close()
        return

    try:
        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=True)
            context = await browser.new_context(viewport={'width': 1920, 'height': 1080})
            page = await context.new_page()

            for i, account in enumerate(accounts):
                if active_credits_jobs.get(job_id, {}).get("status") != "running":
                    break

                email = account["email"]
                active_credits_jobs[job_id]["current_email"] = email

                credits = await _scrape_credits(page, email, PASSWORD)

                if credits:
                    sync_db.login_results.update_one(
                        {"email": email},
                        {"$set": {"credits": credits, "credits_updated_at": datetime.now(timezone.utc).isoformat()}}
                    )
                    active_credits_jobs[job_id]["credits_found"] += 1
                    logger.info(f"Credits for {email}: {credits}")

                active_credits_jobs[job_id]["processed"] = i + 1

            await context.close()
            await browser.close()

    except Exception as e:
        logger.error(f"Credits scan error: {e}")
        active_credits_jobs[job_id]["error"] = str(e)
    finally:
        sync_client.close()

    if active_credits_jobs[job_id]["status"] == "running":
        active_credits_jobs[job_id]["status"] = "completed"
    active_credits_jobs[job_id]["completed_at"] = datetime.now(timezone.utc).isoformat()
    logger.info(f"Credits scan {job_id} done: {active_credits_jobs[job_id]['credits_found']}/{active_credits_jobs[job_id]['processed']} found")


def generate_email(prefix, num, pad, domain):
    if pad > 0:
        num_str = str(num).zfill(pad)
    else:
        num_str = str(num)
    return f"{prefix}{num_str}@{domain}"


def _is_veenu_skipped(num):
    for skip_start, skip_end in VEENU_SKIP_RANGES:
        if skip_start <= num <= skip_end:
            return True
    return False


def generate_all_emails():
    emails = []
    for r in EMAIL_RANGES:
        for num in range(r["start"], r["end"] + 1):
            if r["prefix"] == "veenu" and _is_veenu_skipped(num):
                continue
            email = generate_email(r["prefix"], num, r["pad"], r["domain"])
            emails.append({"email": email, "prefix": r["prefix"], "num": num})
    return emails


def count_total_emails():
    total = 0
    for r in EMAIL_RANGES:
        if r["prefix"] == "veenu":
            for num in range(r["start"], r["end"] + 1):
                if not _is_veenu_skipped(num):
                    total += 1
        else:
            total += r["end"] - r["start"] + 1
    return total


def run_scan_worker(job_id, concurrency=5):
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    loop.run_until_complete(_async_scan_worker(job_id, concurrency))
    loop.close()


async def _async_scan_worker(job_id, concurrency):
    os.environ['PLAYWRIGHT_BROWSERS_PATH'] = '/pw-browsers'
    from playwright.async_api import async_playwright

    sync_client = MongoClient(mongo_url)
    sync_db = sync_client[db_name]

    all_emails = generate_all_emails()

    # RESUME: Get already-tested emails from DB and skip them
    already_tested = set()
    for doc in sync_db.login_results.find({}, {"email": 1}):
        already_tested.add(doc["email"])

    remaining_emails = [e for e in all_emails if e["email"] not in already_tested]
    total_remaining = len(remaining_emails)
    total_all = len(all_emails)
    already_done = len(already_tested)

    # Get current success count from DB
    current_success = sync_db.login_results.count_documents({"status": "success"})

    active_jobs[job_id] = {
        "status": "running",
        "total": total_all,
        "already_tested": already_done,
        "tested": already_done,
        "successful": current_success,
        "failed": already_done - current_success,
        "remaining": total_remaining,
        "current_email": "",
        "started_at": datetime.now(timezone.utc).isoformat()
    }

    # Save scan job to DB for persistence
    sync_db.scan_jobs.update_one(
        {"job_id": job_id},
        {"$set": {
            "job_id": job_id,
            "status": "running",
            "total": total_all,
            "already_tested": already_done,
            "remaining": total_remaining,
            "started_at": datetime.now(timezone.utc).isoformat()
        }},
        upsert=True
    )

    logger.info(f"Scan {job_id}: {total_remaining} remaining of {total_all} total ({already_done} already tested, {current_success} successful)")

    if total_remaining == 0:
        active_jobs[job_id]["status"] = "completed"
        active_jobs[job_id]["completed_at"] = datetime.now(timezone.utc).isoformat()
        sync_db.scan_jobs.update_one({"job_id": job_id}, {"$set": {"status": "completed", "completed_at": datetime.now(timezone.utc).isoformat()}})
        sync_client.close()
        return

    try:
        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=True)

            sem = asyncio.Semaphore(concurrency)
            email_idx = [0]

            async def worker(worker_id):
                context = await browser.new_context(viewport={'width': 1920, 'height': 1080})
                page = await context.new_page()

                while True:
                    async with sem:
                        if email_idx[0] >= total_remaining:
                            break
                        if active_jobs.get(job_id, {}).get("status") != "running":
                            break
                        idx = email_idx[0]
                        email_idx[0] += 1

                    item = remaining_emails[idx]
                    email = item["email"]
                    active_jobs[job_id]["current_email"] = email

                    success = await test_login_fast(page, email, PASSWORD)

                    active_jobs[job_id]["tested"] += 1
                    status = "success" if success else "failed"
                    if success:
                        active_jobs[job_id]["successful"] += 1
                        logger.info(f"SUCCESS: {email}")
                    else:
                        active_jobs[job_id]["failed"] += 1

                    active_jobs[job_id]["remaining"] = total_remaining - email_idx[0]

                    sync_db.login_results.update_one(
                        {"email": email},
                        {"$set": {
                            "email": email,
                            "prefix": item["prefix"],
                            "num": item["num"],
                            "status": status,
                            "job_id": job_id,
                            "tested_at": datetime.now(timezone.utc).isoformat()
                        }},
                        upsert=True
                    )

                try:
                    await context.close()
                except Exception:
                    pass

            tasks = [asyncio.create_task(worker(i)) for i in range(concurrency)]
            await asyncio.gather(*tasks, return_exceptions=True)
            await browser.close()

    except Exception as e:
        logger.error(f"Scan error: {e}")
        active_jobs[job_id]["status"] = "error"
        active_jobs[job_id]["error"] = str(e)
    finally:
        sync_client.close()

    if active_jobs[job_id]["status"] == "running":
        active_jobs[job_id]["status"] = "completed"
    active_jobs[job_id]["completed_at"] = datetime.now(timezone.utc).isoformat()

    # Persist final status
    try:
        s_client = MongoClient(mongo_url)
        s_db = s_client[db_name]
        s_db.scan_jobs.update_one(
            {"job_id": job_id},
            {"$set": {
                "status": active_jobs[job_id]["status"],
                "tested": active_jobs[job_id]["tested"],
                "successful": active_jobs[job_id]["successful"],
                "failed": active_jobs[job_id]["failed"],
                "completed_at": active_jobs[job_id].get("completed_at")
            }}
        )
        s_client.close()
    except Exception:
        pass

    logger.info(f"Scan {job_id} done: {active_jobs[job_id]['successful']} successful / {active_jobs[job_id]['tested']} tested")


async def test_login_fast(page, email, password):
    """Fast login test — returns True/False only. No credit scraping."""
    try:
        await page.goto('https://www.doubledowncasino.com', wait_until='domcontentloaded', timeout=20000)
        await page.wait_for_timeout(2500)

        play_btn = await page.query_selector('img[src*="playnow"]')
        if play_btn:
            await play_btn.click(force=True)
            await page.wait_for_timeout(2000)

        try:
            await page.click('img[src*="email_connect"]', force=True, timeout=5000)
            await page.wait_for_timeout(800)
        except Exception:
            return False

        try:
            await page.click('img[src*="email_dailog_login"]', force=True, timeout=3000)
            await page.wait_for_timeout(400)
        except Exception:
            pass

        try:
            await page.fill('#emailID', email, timeout=3000)
            await page.fill('#pw', password, timeout=3000)
        except Exception:
            return False

        auth_result = {"success": None}
        auth_event = asyncio.Event()

        async def handle_response(response):
            if 'authenticate/user' in response.url:
                try:
                    if response.status == 200:
                        auth_result["success"] = True
                    else:
                        auth_result["success"] = False
                except Exception:
                    auth_result["success"] = False
                auth_event.set()

        page.on('response', handle_response)

        try:
            await page.click('img[src*="green_login"]', force=True, timeout=3000)
        except Exception:
            page.remove_listener('response', handle_response)
            return False

        try:
            await asyncio.wait_for(auth_event.wait(), timeout=12.0)
        except asyncio.TimeoutError:
            auth_result["success"] = False

        page.remove_listener('response', handle_response)
        return auth_result["success"] == True

    except Exception as e:
        logger.debug(f"Login test error for {email}: {e}")
        return False


# ============== API Endpoints ==============
class ScanRequest(BaseModel):
    batch_size: int = 5
    prefixes: Optional[List[str]] = None


@router.post("/start")
async def start_scan(request: ScanRequest = ScanRequest()):
    # Check if a scan is already running in memory
    for jid, job in active_jobs.items():
        if job["status"] == "running":
            return {"message": "Scan already running", "job_id": jid, "status": job}

    job_id = str(uuid.uuid4())

    # DO NOT delete previous results — resume from where we left off
    already_tested = await db.login_results.count_documents({})
    total = count_total_emails()

    thread = threading.Thread(target=run_scan_worker, args=(job_id, request.batch_size), daemon=True)
    thread.start()

    return {
        "job_id": job_id,
        "total_emails": total,
        "already_tested": already_tested,
        "remaining": total - already_tested,
        "status": "started"
    }


@router.post("/stop")
async def stop_scan():
    for jid, job in active_jobs.items():
        if job["status"] == "running":
            job["status"] = "stopped"
            return {"message": "Scan stopped", "job_id": jid}
    return {"message": "No active scan to stop"}


@router.get("/status")
async def get_scan_status():
    # If there's an active in-memory job, return it
    if active_jobs:
        latest_job_id = max(active_jobs.keys(), key=lambda k: active_jobs[k].get("started_at", ""))
        job = active_jobs[latest_job_id]
        return {"job_id": latest_job_id, **{k: v for k, v in job.items()}}

    # Otherwise, check DB for persisted stats
    total = count_total_emails()
    tested = await db.login_results.count_documents({})
    successful = await db.login_results.count_documents({"status": "success"})

    # Check for last completed scan job
    last_job = await db.scan_jobs.find_one({}, sort=[("started_at", -1)])

    return {
        "status": "idle" if tested == 0 else "resumable",
        "total": total,
        "tested": tested,
        "successful": successful,
        "failed": tested - successful,
        "remaining": total - tested,
        "last_job": {
            "job_id": last_job["job_id"],
            "status": last_job["status"],
            "started_at": last_job.get("started_at"),
            "completed_at": last_job.get("completed_at")
        } if last_job else None
    }


@router.get("/results")
async def get_results(status_filter: str = Query("success"), limit: int = 5000, sort_by: str = Query("credits")):
    query = {}
    if status_filter != "all":
        query["status"] = status_filter

    # Sort: credits descending (numeric), then by tested_at
    if sort_by == "credits":
        # Custom sort: accounts with numeric credits first (desc), then unknown, then none
        results = await db.login_results.find(query, {"_id": 0}).to_list(limit)
        def credits_sort_key(r):
            c = r.get("credits")
            if c is None or c == "LOGIN_OK_CREDITS_UNKNOWN":
                return -1
            try:
                return int(c)
            except (ValueError, TypeError):
                return 0
        results.sort(key=credits_sort_key, reverse=True)
    else:
        results = await db.login_results.find(query, {"_id": 0}).sort("tested_at", -1).to_list(limit)

    total_success = await db.login_results.count_documents({"status": "success"})
    total_tested = await db.login_results.count_documents({})
    return {"results": results, "total_success": total_success, "total_tested": total_tested}


@router.post("/reset")
async def reset_results():
    """Only use this to deliberately clear all results and start fresh."""
    await db.login_results.delete_many({})
    await db.scan_jobs.delete_many({})
    active_jobs.clear()
    return {"message": "All results cleared. Next scan will start from the beginning."}


@router.get("/download")
async def download_results():
    results = await db.login_results.find({"status": "success"}, {"_id": 0}).sort("email", 1).to_list(10000)
    if not results:
        raise HTTPException(status_code=404, detail="No successful logins found")

    wb = Workbook()
    ws = wb.active
    ws.title = "Successful Logins"
    header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    headers = ['#', 'Email', 'Prefix', 'Number', 'Credits', 'Last Farmed', 'Tested At']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for idx, r in enumerate(results, 2):
        values = [idx - 1, r.get('email', ''), r.get('prefix', ''), r.get('num', ''), r.get('credits', ''), r.get('last_farmed_at', ''), r.get('tested_at', '')]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=idx, column=col, value=v)
            cell.border = thin_border

    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 25
    ws.freeze_panes = 'A2'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"DDC_Successful_Logins_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.get("/autologin")
async def auto_login_ddc(email: str = Query(...)):
    """Serve an HTML page that opens DDC and auto-fills login credentials."""
    from fastapi.responses import HTMLResponse
    import html as html_mod

    safe_email = html_mod.escape(email)
    safe_password = html_mod.escape(PASSWORD)

    page_html = f"""<!DOCTYPE html>
<html><head>
<meta charset="utf-8">
<title>DDC Auto-Login: {safe_email}</title>
<style>
  body {{ font-family: -apple-system, sans-serif; background: #0a0b0d; color: #f0f2f5; display: flex; flex-direction: column; align-items: center; justify-content: center; height: 100vh; margin: 0; }}
  .card {{ background: #12141a; border: 1px solid #2a2f3a; border-radius: 12px; padding: 32px 40px; text-align: center; max-width: 500px; }}
  h1 {{ font-size: 1.4rem; margin-bottom: 8px; }}
  .email {{ color: #22c55e; font-weight: 600; font-size: 1.1rem; }}
  .info {{ color: #8b919e; font-size: 0.85rem; margin: 16px 0; line-height: 1.6; }}
  .cred {{ background: #1a1d24; border: 1px solid #2a2f3a; border-radius: 8px; padding: 12px 16px; margin: 12px 0; text-align: left; font-family: monospace; font-size: 0.9rem; }}
  .cred span {{ color: #8b919e; }}
  .cred strong {{ color: #f0f2f5; }}
  .btn {{ display: inline-block; margin-top: 16px; padding: 12px 28px; background: #22c55e; color: #000; font-weight: 700; border-radius: 8px; text-decoration: none; font-size: 1rem; cursor: pointer; border: none; }}
  .btn:hover {{ background: #16a34a; }}
  .step {{ color: #f59e0b; font-size: 0.8rem; margin-top: 20px; }}
</style>
</head><body>
<div class="card">
  <h1>DoubleDown Casino Auto-Login</h1>
  <p class="email">{safe_email}</p>
  <div class="cred">
    <span>Email:</span> <strong>{safe_email}</strong><br>
    <span>Password:</span> <strong>{safe_password}</strong>
  </div>
  <div class="info">
    Click the button below to open DoubleDown Casino.<br>
    The login form will be auto-filled with your credentials.
  </div>
  <button class="btn" onclick="openDDC()">Open & Auto-Login</button>
  <p class="step" id="status">Ready to launch...</p>
</div>
<script>
  const EMAIL = "{safe_email}";
  const PW = "{safe_password}";

  function openDDC() {{
    document.getElementById('status').textContent = 'Opening DoubleDown Casino...';
    const ddcWindow = window.open('https://www.doubledowncasino.com', '_blank');

    // Poll the DDC window and auto-fill when ready
    let attempts = 0;
    const maxAttempts = 40;
    const interval = setInterval(() => {{
      attempts++;
      if (attempts > maxAttempts) {{
        clearInterval(interval);
        document.getElementById('status').textContent = 'Auto-fill timed out. Please login manually using the credentials above.';
        return;
      }}
      try {{
        // Try to access the DDC window (may be blocked by CORS)
        if (ddcWindow && !ddcWindow.closed) {{
          ddcWindow.postMessage({{ type: 'ddc-autofill', email: EMAIL, password: PW }}, '*');
        }}
      }} catch(e) {{
        // Cross-origin - expected
      }}
      document.getElementById('status').textContent = 'DDC opened in new tab. Use credentials above if auto-fill is blocked by browser security.';
      if (attempts > 3) clearInterval(interval);
    }}, 2000);
  }}
</script>
</body></html>"""

    return HTMLResponse(content=page_html, status_code=200)


@router.get("/ranges")
async def get_email_ranges():
    ranges = []
    for r in EMAIL_RANGES:
        count = r["end"] - r["start"] + 1
        sample_start = generate_email(r["prefix"], r["start"], r["pad"], r["domain"])
        sample_end = generate_email(r["prefix"], r["end"], r["pad"], r["domain"])
        ranges.append({
            "prefix": r["prefix"], "start": r["start"], "end": r["end"],
            "pad": r["pad"], "count": count,
            "sample_start": sample_start, "sample_end": sample_end
        })
    return {"ranges": ranges, "total": count_total_emails()}


# ============== Credits Scan Endpoints ==============
@router.post("/credits/start")
async def start_credits_scan():
    """Start credits scraping for all successful accounts."""
    for jid, job in active_credits_jobs.items():
        if job["status"] == "running":
            return {"message": "Credits scan already running", "job_id": jid, "status": job}

    pending = await db.login_results.count_documents(
        {"status": "success", "$or": [{"credits": {"$exists": False}}, {"credits": None}]}
    )

    job_id = str(uuid.uuid4())
    thread = threading.Thread(target=run_credits_worker, args=(job_id,), daemon=True)
    thread.start()

    return {"job_id": job_id, "pending_accounts": pending, "status": "started"}


@router.post("/credits/stop")
async def stop_credits_scan():
    for jid, job in active_credits_jobs.items():
        if job["status"] == "running":
            job["status"] = "stopped"
            return {"message": "Credits scan stopped", "job_id": jid}
    return {"message": "No active credits scan"}


@router.get("/credits/status")
async def get_credits_status():
    if active_credits_jobs:
        latest = max(active_credits_jobs.keys(), key=lambda k: active_credits_jobs[k].get("started_at", ""))
        return active_credits_jobs[latest]

    total_success = await db.login_results.count_documents({"status": "success"})
    with_credits = await db.login_results.count_documents({
        "status": "success",
        "credits": {"$exists": True, "$ne": None, "$ne": "LOGIN_OK_CREDITS_UNKNOWN"}
    })
    pending = total_success - with_credits

    return {
        "status": "idle" if pending == 0 else "resumable",
        "total": total_success,
        "with_credits": with_credits,
        "pending": pending
    }


# ============== Chip Farmer ==============
active_farm_jobs = {}

PROMO_CODE_SOURCES = [
    "https://gamehunters.club/doubledown-casino-free-slots/share-links",
    "https://www.giftseize.io/games/doubledown-casino-promo-codes-free-chips",
    "https://thegamereward.com/double-down-casino-codes/",
]


async def _farm_single_account(page, email, password, promo_codes=None):
    """Login to DDC, collect all free bonuses via API interception. Returns dict with results."""
    result = {"chips_before": None, "chips_after": None, "bonuses_collected": [], "errors": []}

    try:
        session_info = {"session_id": None, "user_id": None, "base_url": None}
        credits_captured = {"before": None, "after": None}

        async def capture_all(response):
            url = response.url
            try:
                if 'lobby/game' in url:
                    text = await response.text()
                    import re as _re
                    cash_match = _re.search(r'"cash"\s*:\s*(\d+)', text)
                    session_match = _re.search(r'"sessionId"\s*:\s*"([^"]+)"', text)
                    if cash_match:
                        if credits_captured["before"] is None:
                            credits_captured["before"] = cash_match.group(1)
                        credits_captured["after"] = cash_match.group(1)
                    if session_match:
                        session_info["session_id"] = session_match.group(1)
                    # Extract base URL for httpbox calls
                    if 'doubledowncasino2.com' in url:
                        session_info["base_url"] = url.split('/v2/lobby')[0]

                elif 'lobby/meta' in url:
                    text = await response.text()
                    if 'rewardAvailable' in text:
                        result["bonuses_collected"].append("meta_loaded")

                elif 'httpbox/poll' in url:
                    text = await response.text()
                    if 'rewardAvailable' in text and 'true' in text:
                        result["bonuses_collected"].append("reward_available")
            except Exception:
                pass

        page.on('response', capture_all)

        # Login flow
        await page.goto('https://www.doubledowncasino.com', wait_until='domcontentloaded', timeout=20000)
        await page.wait_for_timeout(2500)

        play_btn = await page.query_selector('img[src*="playnow"]')
        if play_btn:
            await play_btn.click(force=True)
            await page.wait_for_timeout(2000)

        try:
            await page.click('img[src*="email_connect"]', force=True, timeout=5000)
            await page.wait_for_timeout(800)
        except Exception:
            page.remove_listener('response', capture_all)
            result["errors"].append("no_email_btn")
            return result

        try:
            await page.click('img[src*="email_dailog_login"]', force=True, timeout=3000)
            await page.wait_for_timeout(400)
        except Exception:
            pass

        try:
            await page.fill('#emailID', email, timeout=3000)
            await page.fill('#pw', password, timeout=3000)
        except Exception:
            page.remove_listener('response', capture_all)
            result["errors"].append("no_form")
            return result

        try:
            await page.click('img[src*="green_login"]', force=True, timeout=3000)
        except Exception:
            page.remove_listener('response', capture_all)
            result["errors"].append("no_login_btn")
            return result

        # Wait for game lobby to load
        await page.wait_for_timeout(12000)
        result["chips_before"] = credits_captured["before"]

        # Now try to claim rewards via the game's API by executing JS in page context
        if session_info["session_id"] and session_info["base_url"]:
            base = session_info["base_url"]
            sid = session_info["session_id"]

            # Try claiming league reward
            try:
                league_result = await page.evaluate(f"""
                    async () => {{
                        try {{
                            const resp = await fetch('{base}/v3/league/reward', {{
                                method: 'POST',
                                headers: {{'Content-Type': 'application/x-www-form-urlencoded'}},
                                body: 'xrid=' + Date.now()
                            }});
                            const text = await resp.text();
                            return text.substring(0, 300);
                        }} catch(e) {{ return 'error:' + e.message; }}
                    }}
                """)
                if league_result and 'error' not in str(league_result).lower():
                    result["bonuses_collected"].append(f"league_reward")
                    logger.info(f"  League reward for {email}: {str(league_result)[:100]}")
            except Exception:
                pass

            # Try claiming daily bonus / wheel spin
            try:
                daily_result = await page.evaluate(f"""
                    async () => {{
                        try {{
                            const resp = await fetch('{base}/v2/daily/bonus', {{
                                method: 'POST',
                                headers: {{'Content-Type': 'application/x-www-form-urlencoded'}},
                                body: 'xrid=' + Date.now()
                            }});
                            return await resp.text();
                        }} catch(e) {{ return 'error:' + e.message; }}
                    }}
                """)
                if daily_result and 'error' not in str(daily_result).lower():
                    result["bonuses_collected"].append("daily_bonus")
            except Exception:
                pass

            # Try SFS httpbox command for collecting gifts
            try:
                gift_cmd = '{"t":"xt","b":{"c":"collectGift","r":-1,"x":"casinoExtension","p":{}}}'
                gift_result = await page.evaluate(f"""
                    async () => {{
                        try {{
                            const resp = await fetch('{base}/httpbox/poll', {{
                                method: 'POST',
                                headers: {{'Content-Type': 'application/x-www-form-urlencoded'}},
                                body: 'sfsHttp={sid}' + encodeURIComponent('{gift_cmd}') + '&ts=' + Date.now() + '&prevts=' + (Date.now()-20000) + '&xrid=' + Date.now()
                            }});
                            return await resp.text();
                        }} catch(e) {{ return 'error:' + e.message; }}
                    }}
                """)
                if gift_result and 'error' not in str(gift_result).lower():
                    result["bonuses_collected"].append("gift_collect")
            except Exception:
                pass

            # Try claiming wheel spin
            try:
                wheel_cmd = '{"t":"xt","b":{"c":"spinWheel","r":-1,"x":"casinoExtension","p":{}}}'
                wheel_result = await page.evaluate(f"""
                    async () => {{
                        try {{
                            const resp = await fetch('{base}/httpbox/poll', {{
                                method: 'POST',
                                headers: {{'Content-Type': 'application/x-www-form-urlencoded'}},
                                body: 'sfsHttp={sid}' + encodeURIComponent('{wheel_cmd}') + '&ts=' + Date.now() + '&prevts=' + (Date.now()-20000) + '&xrid=' + Date.now()
                            }});
                            return await resp.text();
                        }} catch(e) {{ return 'error:' + e.message; }}
                    }}
                """)
                if wheel_result and 'error' not in str(wheel_result).lower():
                    result["bonuses_collected"].append("wheel_spin")
            except Exception:
                pass

        # Wait and re-check balance
        await page.wait_for_timeout(5000)

        # Get final balance by reloading
        credits_captured["after"] = None
        try:
            reload_result = await page.evaluate(f"""
                async () => {{
                    try {{
                        const resp = await fetch('{session_info["base_url"]}/v2/lobby/game', {{
                            method: 'POST',
                            headers: {{'Content-Type': 'application/x-www-form-urlencoded'}},
                            body: 'language=en&xrid=' + Date.now()
                        }});
                        return await resp.text();
                    }} catch(e) {{ return ''; }}
                }}
            """) if session_info["base_url"] else ""
            import re as _re
            cash_match = _re.search(r'"cash"\s*:\s*(\d+)', reload_result or "")
            if cash_match:
                credits_captured["after"] = cash_match.group(1)
        except Exception:
            pass

        result["chips_after"] = credits_captured["after"] or credits_captured["before"]
        page.remove_listener('response', capture_all)

    except Exception as e:
        result["errors"].append(str(e))

    return result


def _fetch_promo_codes():
    """Scrape latest promo codes from collector sites."""
    import requests
    import re
    codes = []
    try:
        # Try gamehunters.club for share links
        resp = requests.get(PROMO_CODE_SOURCES[0], timeout=15, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
        })
        if resp.status_code == 200:
            # Extract share link IDs
            links = re.findall(r'share-links/click/(\d+)', resp.text)
            for link_id in links[:10]:
                codes.append(f"https://gamehunters.club/doubledown-casino-free-slots/share-links/click/{link_id}")
    except Exception as e:
        logger.debug(f"Promo code fetch error: {e}")

    try:
        # Try giftseize for actual codes
        resp = requests.get(PROMO_CODE_SOURCES[1], timeout=15, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
        })
        if resp.status_code == 200:
            import re
            code_matches = re.findall(r'[A-Z0-9]{6,20}', resp.text)
            for c in code_matches[:10]:
                if len(c) >= 8 and not c.startswith('HTTP'):
                    codes.append(c)
    except Exception:
        pass

    return codes


def run_farm_worker(job_id):
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    loop.run_until_complete(_async_farm_worker(job_id))
    loop.close()


async def _async_farm_worker(job_id):
    """Farm chips for all successful accounts."""
    os.environ['PLAYWRIGHT_BROWSERS_PATH'] = '/pw-browsers'
    from playwright.async_api import async_playwright

    sync_client = MongoClient(mongo_url)
    sync_db = sync_client[db_name]

    accounts = list(sync_db.login_results.find({"status": "success"}, {"email": 1}))
    total = len(accounts)

    # Fetch promo codes
    promo_codes = _fetch_promo_codes()
    logger.info(f"Chip Farm: {total} accounts, {len(promo_codes)} promo codes found")

    active_farm_jobs[job_id] = {
        "status": "running",
        "total": total,
        "processed": 0,
        "chips_gained": 0,
        "current_email": "",
        "promo_codes_found": len(promo_codes),
        "started_at": datetime.now(timezone.utc).isoformat()
    }

    try:
        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=True)
            context = await browser.new_context(viewport={'width': 1920, 'height': 1080})
            page = await context.new_page()

            for i, account in enumerate(accounts):
                if active_farm_jobs.get(job_id, {}).get("status") != "running":
                    break

                email = account["email"]
                active_farm_jobs[job_id]["current_email"] = email

                farm_result = await _farm_single_account(page, email, PASSWORD, promo_codes)

                # Update DB with farming results
                update = {
                    "last_farmed_at": datetime.now(timezone.utc).isoformat(),
                    "farm_bonuses": farm_result["bonuses_collected"],
                    "farm_errors": farm_result["errors"]
                }
                if farm_result["chips_before"]:
                    update["credits_before_farm"] = farm_result["chips_before"]
                if farm_result["chips_after"]:
                    update["credits"] = farm_result["chips_after"]
                    update["credits_after_farm"] = farm_result["chips_after"]
                    update["credits_updated_at"] = datetime.now(timezone.utc).isoformat()
                elif farm_result["chips_before"]:
                    update["credits"] = farm_result["chips_before"]
                    update["credits_updated_at"] = datetime.now(timezone.utc).isoformat()

                # Calculate gain
                before = int(farm_result["chips_before"] or 0)
                after = int(farm_result["chips_after"] or farm_result["chips_before"] or 0)
                if after > before:
                    active_farm_jobs[job_id]["chips_gained"] += (after - before)

                sync_db.login_results.update_one({"email": email}, {"$set": update})
                active_farm_jobs[job_id]["processed"] = i + 1

                if farm_result["chips_after"]:
                    logger.info(f"Farmed {email}: {farm_result['chips_before']} → {farm_result['chips_after']} | Bonuses: {farm_result['bonuses_collected']}")

            await context.close()
            await browser.close()

    except Exception as e:
        logger.error(f"Farm error: {e}")
        active_farm_jobs[job_id]["error"] = str(e)
    finally:
        sync_client.close()

    if active_farm_jobs[job_id]["status"] == "running":
        active_farm_jobs[job_id]["status"] = "completed"
    active_farm_jobs[job_id]["completed_at"] = datetime.now(timezone.utc).isoformat()


@router.post("/farm/start")
async def start_chip_farm():
    for jid, job in active_farm_jobs.items():
        if job["status"] == "running":
            return {"message": "Farm already running", "job_id": jid, "status": job}

    total = await db.login_results.count_documents({"status": "success"})
    job_id = str(uuid.uuid4())
    thread = threading.Thread(target=run_farm_worker, args=(job_id,), daemon=True)
    thread.start()
    return {"job_id": job_id, "total_accounts": total, "status": "started"}


@router.post("/farm/stop")
async def stop_chip_farm():
    for jid, job in active_farm_jobs.items():
        if job["status"] == "running":
            job["status"] = "stopped"
            return {"message": "Farm stopped", "job_id": jid}
    return {"message": "No active farm"}


@router.get("/farm/status")
async def get_farm_status():
    if active_farm_jobs:
        latest = max(active_farm_jobs.keys(), key=lambda k: active_farm_jobs[k].get("started_at", ""))
        return active_farm_jobs[latest]
    return {"status": "idle"}
