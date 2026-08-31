"""
LinkedLeads.ai — Extension API
Backend endpoints for the Chrome Extension to:
- Report LinkedIn session status
- Poll for pending campaign tasks
- Report task execution results
- Get dashboard stats
"""
import os
import logging
from datetime import datetime, timezone, timedelta
from fastapi import APIRouter, HTTPException, Request
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel
from typing import Optional
from routes.crm_auth import get_current_user

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/ext", tags=["extension"])

mongo_url = os.environ.get("MONGO_URL", "")
_client = AsyncIOMotorClient(mongo_url)
_db = _client[os.environ.get("DB_NAME", "test_database")]


# ============ Models ============
class SessionReport(BaseModel):
    active: bool
    li_at_prefix: Optional[str] = None


class TaskResult(BaseModel):
    success: bool
    error: Optional[str] = None
    action: Optional[str] = None
    note: Optional[str] = None
    publicId: Optional[str] = None


# ============ Session ============
@router.post("/session")
async def report_session(body: SessionReport, request: Request):
    user = await get_current_user(request)
    await _db.ext_sessions.update_one(
        {"user_id": user["id"]},
        {"$set": {
            "user_id": user["id"],
            "email": user["email"],
            "active": body.active,
            "li_at_prefix": body.li_at_prefix,
            "last_seen": datetime.now(timezone.utc).isoformat(),
        }},
        upsert=True,
    )
    return {"ok": True}


# ============ Task Polling ============
@router.get("/tasks/next")
async def get_next_task(request: Request):
    """Get the next pending task for this user's active campaigns."""
    user = await get_current_user(request)

    # Check daily limit
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    today_count = await _db.ext_task_log.count_documents({
        "user_id": user["id"],
        "date": today,
        "success": True,
    })

    # Get user's daily limit from settings or default
    user_doc = await _db.crm_users.find_one({"email": user["email"]})
    daily_limit = (user_doc or {}).get("daily_limit", 20)

    if today_count >= daily_limit:
        return {"task": None, "reason": "daily_limit_reached", "count": today_count, "limit": daily_limit}

    # Find next pending task from active campaigns
    task = await _db.ext_tasks.find_one(
        {
            "user_id": user["id"],
            "status": "pending",
        },
        sort=[("priority", -1), ("created_at", 1)],
    )

    if not task:
        return {"task": None, "reason": "no_pending_tasks"}

    # Mark as in_progress
    await _db.ext_tasks.update_one(
        {"_id": task["_id"]},
        {"$set": {"status": "in_progress", "started_at": datetime.now(timezone.utc).isoformat()}}
    )

    return {
        "task": {
            "task_id": str(task["_id"]),
            "campaign_id": task.get("campaign_id", ""),
            "type": task["type"],
            "target_profile_url": task.get("target_profile_url", ""),
            "target_public_id": task.get("target_public_id", ""),
            "message": task.get("message", ""),
            "prospect": task.get("prospect", {}),
        }
    }


# ============ Task Result ============
@router.post("/tasks/{task_id}/result")
async def report_task_result(task_id: str, body: TaskResult, request: Request):
    user = await get_current_user(request)
    from bson import ObjectId
    from bson.errors import InvalidId

    try:
        oid = ObjectId(task_id)
    except (InvalidId, Exception):
        raise HTTPException(status_code=400, detail="Invalid task ID format")

    task = await _db.ext_tasks.find_one({"_id": oid, "user_id": user["id"]})
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")

    new_status = "completed" if body.success else "failed"
    await _db.ext_tasks.update_one(
        {"_id": oid},
        {"$set": {
            "status": new_status,
            "result": body.dict(),
            "completed_at": datetime.now(timezone.utc).isoformat(),
        }}
    )

    # Log for daily tracking
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    await _db.ext_task_log.insert_one({
        "user_id": user["id"],
        "task_id": task_id,
        "campaign_id": task.get("campaign_id", ""),
        "type": task["type"],
        "target_public_id": task.get("target_public_id", ""),
        "success": body.success,
        "error": body.error,
        "date": today,
        "timestamp": datetime.now(timezone.utc).isoformat(),
    })

    # Update campaign stats
    if task.get("campaign_id"):
        if body.success:
            await _db.ext_campaigns.update_one(
                {"campaign_id": task["campaign_id"]},
                {"$inc": {"completed_count": 1}}
            )
        else:
            await _db.ext_campaigns.update_one(
                {"campaign_id": task["campaign_id"]},
                {"$inc": {"failed_count": 1}}
            )

    return {"ok": True, "status": new_status}


# ============ Dashboard Stats ============
@router.get("/stats")
async def get_stats(request: Request):
    user = await get_current_user(request)
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    # Today's stats
    pipeline = [
        {"$match": {"user_id": user["id"], "date": today, "success": True}},
        {"$group": {"_id": "$type", "count": {"$sum": 1}}},
    ]
    results = await _db.ext_task_log.aggregate(pipeline).to_list(10)
    stats = {r["_id"]: r["count"] for r in results}

    # Active campaigns count
    active_campaigns = await _db.ext_campaigns.count_documents({
        "user_id": user["id"],
        "status": {"$in": ["active", "running"]},
    })

    # Pending tasks
    pending_tasks = await _db.ext_tasks.count_documents({
        "user_id": user["id"],
        "status": "pending",
    })

    # Total all-time
    total_completed = await _db.ext_task_log.count_documents({
        "user_id": user["id"],
        "success": True,
    })

    return {
        "today": {
            "connects": stats.get("connect", 0),
            "messages": stats.get("message", 0),
            "visits": stats.get("visit", 0),
        },
        "active_campaigns": active_campaigns,
        "pending_tasks": pending_tasks,
        "total_completed": total_completed,
    }


# ============ Campaigns CRUD ============
class CreateCampaignRequest(BaseModel):
    name: str
    type: str = "connect"  # connect, message, sequence
    message_template: Optional[str] = ""
    daily_limit: int = 20
    prospects: list = []  # [{profile_url, public_id, name, company, title}]


@router.get("/campaigns")
async def list_campaigns(request: Request):
    user = await get_current_user(request)
    campaigns = await _db.ext_campaigns.find(
        {"user_id": user["id"]},
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    return {"campaigns": campaigns}


@router.post("/campaigns")
async def create_campaign(body: CreateCampaignRequest, request: Request):
    user = await get_current_user(request)
    import uuid

    campaign_id = str(uuid.uuid4())[:12]
    now = datetime.now(timezone.utc).isoformat()

    campaign_doc = {
        "campaign_id": campaign_id,
        "user_id": user["id"],
        "user_email": user["email"],
        "name": body.name,
        "type": body.type,
        "message_template": body.message_template,
        "daily_limit": body.daily_limit,
        "status": "active",
        "total_prospects": len(body.prospects),
        "completed_count": 0,
        "failed_count": 0,
        "created_at": now,
        "updated_at": now,
    }
    await _db.ext_campaigns.insert_one(campaign_doc)

    # Create tasks for each prospect
    tasks = []
    for i, p in enumerate(body.prospects):
        public_id = p.get("public_id") or _extract_public_id(p.get("profile_url", ""))
        tasks.append({
            "campaign_id": campaign_id,
            "user_id": user["id"],
            "type": body.type,
            "target_profile_url": p.get("profile_url", ""),
            "target_public_id": public_id,
            "message": body.message_template,
            "prospect": {
                "name": p.get("name", ""),
                "company": p.get("company", ""),
                "title": p.get("title", ""),
                "location": p.get("location", ""),
            },
            "status": "pending",
            "priority": 0,
            "created_at": now,
            "order": i,
        })

    if tasks:
        await _db.ext_tasks.insert_many(tasks)

    return {
        "campaign_id": campaign_id,
        "tasks_created": len(tasks),
        "status": "active",
    }


@router.post("/campaigns/{campaign_id}/pause")
async def pause_campaign(campaign_id: str, request: Request):
    user = await get_current_user(request)
    await _db.ext_campaigns.update_one(
        {"campaign_id": campaign_id, "user_id": user["id"]},
        {"$set": {"status": "paused", "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    # Pause pending tasks
    await _db.ext_tasks.update_many(
        {"campaign_id": campaign_id, "user_id": user["id"], "status": "pending"},
        {"$set": {"status": "paused"}}
    )
    return {"ok": True}


@router.post("/campaigns/{campaign_id}/resume")
async def resume_campaign(campaign_id: str, request: Request):
    user = await get_current_user(request)
    await _db.ext_campaigns.update_one(
        {"campaign_id": campaign_id, "user_id": user["id"]},
        {"$set": {"status": "active", "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    await _db.ext_tasks.update_many(
        {"campaign_id": campaign_id, "user_id": user["id"], "status": "paused"},
        {"$set": {"status": "pending"}}
    )
    return {"ok": True}


@router.delete("/campaigns/{campaign_id}")
async def delete_campaign(campaign_id: str, request: Request):
    user = await get_current_user(request)
    await _db.ext_campaigns.delete_one({"campaign_id": campaign_id, "user_id": user["id"]})
    await _db.ext_tasks.delete_many({"campaign_id": campaign_id, "user_id": user["id"]})
    return {"ok": True}


# ============ Upload Prospects (XLSX) ============
from fastapi import UploadFile, File, Form
import openpyxl
import io


@router.post("/campaigns/{campaign_id}/upload-prospects")
async def upload_prospects(campaign_id: str, request: Request, file: UploadFile = File(...)):
    user = await get_current_user(request)

    campaign = await _db.ext_campaigns.find_one({"campaign_id": campaign_id, "user_id": user["id"]})
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid spreadsheet file. Please upload a valid .xlsx file.")
    ws = wb.active

    try:
        headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    except (StopIteration, Exception):
        wb.close()
        raise HTTPException(status_code=400, detail="Spreadsheet is empty or has no headers.")

    # Map columns
    col_map = {}
    for i, h in enumerate(headers):
        if "url" in h or "linkedin" in h or "profile" in h:
            col_map["profile_url"] = i
        elif "name" in h:
            col_map["name"] = i
        elif "company" in h or "org" in h:
            col_map["company"] = i
        elif "title" in h or "position" in h or "role" in h:
            col_map["title"] = i
        elif "location" in h or "city" in h:
            col_map["location"] = i

    tasks = []
    now = datetime.now(timezone.utc).isoformat()
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        cells = list(row)
        profile_url = str(cells[col_map.get("profile_url", 0)] or "").strip()
        if not profile_url:
            continue
        public_id = _extract_public_id(profile_url)
        name = str(cells[col_map.get("name", 1)] or "").strip() if col_map.get("name") is not None and col_map["name"] < len(cells) else ""
        company = str(cells[col_map.get("company", 2)] or "").strip() if col_map.get("company") is not None and col_map["company"] < len(cells) else ""
        title = str(cells[col_map.get("title", 3)] or "").strip() if col_map.get("title") is not None and col_map["title"] < len(cells) else ""
        location = str(cells[col_map.get("location")] or "").strip() if col_map.get("location") is not None and col_map["location"] < len(cells) else ""

        tasks.append({
            "campaign_id": campaign_id,
            "user_id": user["id"],
            "type": campaign.get("type", "connect"),
            "target_profile_url": profile_url,
            "target_public_id": public_id,
            "message": campaign.get("message_template", ""),
            "prospect": {"name": name, "company": company, "title": title, "location": location},
            "status": "pending",
            "priority": 0,
            "created_at": now,
            "order": count,
        })
        count += 1

    if tasks:
        await _db.ext_tasks.insert_many(tasks)
        await _db.ext_campaigns.update_one(
            {"campaign_id": campaign_id},
            {"$inc": {"total_prospects": len(tasks)}, "$set": {"updated_at": now}}
        )

    wb.close()
    return {"prospects_added": len(tasks), "campaign_id": campaign_id}


# ============ Campaign Tasks List ============
@router.get("/campaigns/{campaign_id}/tasks")
async def get_campaign_tasks(campaign_id: str, request: Request, status: str = "all"):
    user = await get_current_user(request)
    query = {"campaign_id": campaign_id, "user_id": user["id"]}
    if status != "all":
        query["status"] = status

    tasks = await _db.ext_tasks.find(query).sort("order", 1).to_list(500)
    result = []
    for t in tasks:
        result.append({
            "task_id": str(t["_id"]),
            "type": t["type"],
            "status": t["status"],
            "target_public_id": t.get("target_public_id", ""),
            "target_profile_url": t.get("target_profile_url", ""),
            "prospect": t.get("prospect", {}),
            "result": t.get("result"),
            "created_at": t.get("created_at"),
            "completed_at": t.get("completed_at"),
        })
    return {"tasks": result, "total": len(result)}


# ============ Helpers ============
def _extract_public_id(url: str) -> str:
    if not url:
        return ""
    import re
    m = re.search(r'linkedin\.com/in/([^/?#]+)', url)
    return m.group(1) if m else ""


# ============ Extension Download ============
@router.get("/download")
async def download_extension():
    """Package and serve the Chrome Extension as a ZIP file."""
    import zipfile
    import io as iomod
    from fastapi.responses import Response

    ext_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "chrome-extension-dist")
    if not os.path.exists(ext_dir):
        ext_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "chrome-extension")

    if not os.path.exists(ext_dir):
        raise HTTPException(status_code=404, detail="Extension files not found")

    buf = iomod.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for root, dirs, files in os.walk(ext_dir):
            for fname in files:
                fpath = os.path.join(root, fname)
                arcname = os.path.relpath(fpath, ext_dir)
                zf.write(fpath, arcname)

    buf.seek(0)
    return Response(
        content=buf.read(),
        media_type="application/zip",
        headers={"Content-Disposition": 'attachment; filename="linkedleads-chrome-extension.zip"'}
    )
