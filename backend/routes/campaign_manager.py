"""
LinkedIn Outreach Campaign Manager v2
Fixed: background task sending, atomic claiming, retry mechanism, cookie preflight.
Uses Emergent Object Storage for persistent file hosting across deployments.
"""
import os
import uuid
import logging
import asyncio
import openpyxl
from pathlib import Path
from datetime import datetime, timezone
from fastapi import APIRouter, HTTPException
from motor.motor_asyncio import AsyncIOMotorClient
import httpx

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/campaigns", tags=["campaigns"])

mongo_url = os.environ.get('MONGO_URL', '')
_client = AsyncIOMotorClient(mongo_url)
_db = _client[os.environ.get('DB_NAME', 'test_database')]

MAX_MESSAGES_PER_DAY = 25

PROSPECTS_FILE = Path(__file__).parent.parent / "uploads" / "prospects.xlsx"
BROCHURE_FILE = Path(__file__).parent.parent / "uploads" / "fundle_marketplace_brochure.pdf"

# Cloud storage paths
CLOUD_PROSPECTS_PATH = "vnagents-crm/campaigns/prospects.xlsx"
CLOUD_BROCHURE_PATH = "vnagents-crm/campaigns/fundle_marketplace_brochure.pdf"

MYNTRA_CAMPAIGN_TEMPLATE = {
    "campaign_id": "7701ea79",
    "name": "Myntra 500 - Marketplace AutoPilot Outreach",
    "message_template": "Dear {name},\n\nHope you are doing well! I was keen to showcase and demo our Marketplace Automation Platform — Marketplace AutoPilot — to your team at {brand}.\n\nIt helps manage sales, commissions, reconciliations, discrepancy detection, and invoicing with leading marketplaces like Myntra, Ajio, Flipkart, Amazon, and Nykaa.\n\nSince {brand} sells on these marketplaces, I thought it would be great for you to see our platform in action.\n\nAttaching a quick e-brochure for you to understand the functionality and how it can help your backend marketplace operations.\n\nWould be happy to have the team do a detailed demo as required.\n\nBrochure: https://kazob2b.fundlezone.com/brochure\n\nRegards,\nVineet Narang\nFundle.ai\nWhatsApp: +91-9910530372",
    "attachment_cloud_path": CLOUD_BROCHURE_PATH,
    "sender_name": "Vineet",
    "daily_limit": 25,
    "status": "active",
}


def _ensure_prospects_file() -> str:
    """Get a path to prospects.xlsx — local first, cloud fallback."""
    if PROSPECTS_FILE.exists():
        return str(PROSPECTS_FILE)
    try:
        from utils.object_storage import download_to_local
        local = str(PROSPECTS_FILE)
        download_to_local(CLOUD_PROSPECTS_PATH, local)
        logger.info("Prospects downloaded from cloud storage")
        return local
    except Exception as e:
        logger.warning(f"Could not download prospects from cloud: {e}")
        return None


async def seed_myntra_campaign():
    """Auto-seed the Myntra 500 campaign if it doesn't exist. Also uploads assets to cloud."""
    try:
        # Upload assets to cloud (idempotent — overwrites are fine)
        try:
            await asyncio.to_thread(_upload_campaign_assets_inner)
        except Exception as e:
            logger.warning(f"Asset upload failed (non-fatal): {e}")

        existing = await _db.outreach_campaigns.find_one({"campaign_id": "7701ea79"})
        if existing:
            # Ensure cloud path is set on existing doc
            if not existing.get("attachment_cloud_path"):
                await _db.outreach_campaigns.update_one(
                    {"campaign_id": "7701ea79"},
                    {"$set": {"attachment_cloud_path": CLOUD_BROCHURE_PATH}}
                )
            return  # Already seeded

        xlsx_path = _ensure_prospects_file()
        if not xlsx_path:
            logger.warning("Myntra prospects file not found locally or in cloud, skipping seed")
            return

        # Create campaign
        doc = {**MYNTRA_CAMPAIGN_TEMPLATE, "created_at": datetime.now(timezone.utc).isoformat()}
        await _db.outreach_campaigns.insert_one(doc)
        logger.info("Myntra 500 campaign created")

        # Import prospects
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        imported = 0
        for row in ws.iter_rows(min_row=5, values_only=True):
            if not row[0] or not str(row[0]).isdigit():
                continue
            linkedin_url = row[16] or ""
            if not linkedin_url or "linkedin.com" not in str(linkedin_url):
                continue
            public_id = str(linkedin_url).rstrip("/").split("/")[-1]
            existing_p = await _db.campaign_prospects.find_one({"campaign_id": "7701ea79", "public_id": public_id})
            if existing_p:
                continue
            await _db.campaign_prospects.insert_one({
                "campaign_id": "7701ea79", "rank": int(row[0]),
                "brand": row[1] or "", "name": row[17] or "", "title": row[21] or "",
                "category": row[2] or "", "pitch": row[13] or "",
                "linkedin_url": str(linkedin_url), "public_id": public_id,
                "status": "pending", "created_at": datetime.now(timezone.utc).isoformat(),
                "sent_at": None, "error": None,
            })
            imported += 1
        logger.info(f"Myntra 500: imported {imported} prospects")
    except Exception as e:
        logger.error(f"Myntra campaign seed error: {e}")


def _upload_campaign_assets_inner():
    """Upload campaign assets to cloud storage (sync, called via asyncio.to_thread)."""
    try:
        from utils.object_storage import init_storage, upload_local_file
        init_storage()

        if BROCHURE_FILE.exists():
            result = upload_local_file(str(BROCHURE_FILE), CLOUD_BROCHURE_PATH, "application/pdf")
            logger.info(f"Brochure uploaded to cloud: {result.get('path')} ({result.get('size')} bytes)")

        if PROSPECTS_FILE.exists():
            result = upload_local_file(str(PROSPECTS_FILE), CLOUD_PROSPECTS_PATH, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            logger.info(f"Prospects uploaded to cloud: {result.get('path')} ({result.get('size')} bytes)")
    except Exception as e:
        logger.error(f"Campaign asset upload error (non-fatal): {e}")



def _build_voyager_headers(li_at: str, jsessionid: str) -> dict:
    clean_js = jsessionid.strip('"').replace("ajax:", "")
    return {
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'accept': 'application/vnd.linkedin.normalized+json+2.1',
        'x-restli-protocol-version': '2.0.0',
        'content-type': 'application/json',
        'cookie': f'li_at={li_at}; JSESSIONID="ajax:{clean_js}"',
        'csrf-token': f'ajax:{clean_js}',
    }


@router.get("/cloud-assets/status")
async def cloud_assets_status():
    """Check if campaign assets are available in cloud storage."""
    results = {"brochure": False, "prospects": False}
    try:
        from utils.object_storage import init_storage, get_object
        init_storage()
        try:
            get_object(CLOUD_BROCHURE_PATH)
            results["brochure"] = True
        except Exception:
            pass
        try:
            get_object(CLOUD_PROSPECTS_PATH)
            results["prospects"] = True
        except Exception:
            pass
    except Exception as e:
        logger.error(f"Cloud asset status check error: {e}")
    return results


@router.post("/cloud-assets/upload")
async def trigger_asset_upload():
    """Manually trigger upload of local campaign files to cloud storage."""
    try:
        _upload_campaign_assets_inner()
        return {"success": True, "detail": "Assets uploaded to cloud storage"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


async def _validate_cookie():
    """Pre-flight cookie validation. Returns (headers, error_msg)."""
    cookie_doc = await _db.li_search_config.find_one({"type": "cookie"})
    if not cookie_doc or not cookie_doc.get("li_at"):
        return None, "No LinkedIn cookie configured. Go to Settings and add your li_at cookie."
    li_at = cookie_doc["li_at"]
    jsessionid = cookie_doc.get("jsessionid", "")
    if not jsessionid:
        return None, "No JSESSIONID. Update cookie in Settings."
    return _build_voyager_headers(li_at, jsessionid), None


async def _resolve_member_urn(http_client, public_id: str, headers: dict):
    """Resolve a LinkedIn public_id to a messaging-compatible URN."""
    resp = await http_client.get(
        f"https://www.linkedin.com/voyager/api/identity/dash/profiles?q=memberIdentity&memberIdentity={public_id}",
        headers=headers,
    )
    if resp.status_code in [401, 403]:
        return None, "cookie_expired"
    if resp.status_code == 302:
        return None, "cookie_rejected"
    if resp.status_code != 200:
        return None, f"profile_lookup_{resp.status_code}"

    data = resp.json()
    # Find the fsd_profile URN (for messaging)
    for item in data.get("included", []):
        eid = item.get("entityUrn", "")
        if "fsd_profile:" in eid:
            return eid, None
        obj_urn = item.get("objectUrn", "")
        if "fsd_profile:" in obj_urn:
            return obj_urn, None
        if "fs_miniProfile:" in eid:
            return eid, None
    return None, "no_urn_found"


async def _send_single_message(http_client, headers: dict, recipient_urn: str, public_id: str, message_text: str):
    """Send a LinkedIn connection request with a personalized note (for non-connections),
    or a direct message (for existing connections)."""

    # Truncate note to 300 chars (LinkedIn connection request limit)
    connection_note = message_text[:300]

    # First try direct message (works for 1st-degree connections)
    msg_payload = {
        "message": {"body": {"text": message_text, "attributes": []}},
        "mailboxUrn": recipient_urn,
        "trackingId": str(uuid.uuid4()),
        "dedupeByClientGeneratedToken": False,
        "hostRecipientUrns": [recipient_urn],
    }
    resp = await http_client.post(
        "https://www.linkedin.com/voyager/api/voyagerMessagingDashMessengerMessages?action=createMessage",
        json=msg_payload, headers=headers,
    )
    if resp.status_code in [200, 201]:
        return True, "direct_message_sent"

    # If messaging fails (403 = not connected), send connection request with note
    invite_payload = {
        "inviteeProfileUrn": recipient_urn,
        "customMessage": connection_note,
        "trackingId": str(uuid.uuid4()),
    }
    resp2 = await http_client.post(
        "https://www.linkedin.com/voyager/api/voyagerRelationshipsDashMemberRelationships?action=verifyQuotaAndCreate",
        json=invite_payload, headers=headers,
    )
    if resp2.status_code in [200, 201]:
        return True, "connection_request_sent"

    return False, resp2.text[:300]


# ============== CRUD ==============
@router.get("")
async def list_campaigns():
    campaigns = []
    async for c in _db.outreach_campaigns.find({}, {"_id": 0}).sort("created_at", -1):
        total = await _db.campaign_prospects.count_documents({"campaign_id": c["campaign_id"]})
        sent = await _db.campaign_prospects.count_documents({"campaign_id": c["campaign_id"], "status": "sent"})
        failed = await _db.campaign_prospects.count_documents({"campaign_id": c["campaign_id"], "status": "failed"})
        pending = await _db.campaign_prospects.count_documents({"campaign_id": c["campaign_id"], "status": {"$in": ["pending", "sending"]}})
        c.update({"total": total, "sent": sent, "failed": failed, "pending": pending})
        campaigns.append(c)
    return {"campaigns": campaigns}


@router.post("")
async def create_campaign(data: dict):
    campaign_id = str(uuid.uuid4())[:8]
    doc = {
        "campaign_id": campaign_id,
        "name": data.get("name", "Untitled Campaign"),
        "message_template": data.get("message_template", ""),
        "attachment_path": data.get("attachment_path", ""),
        "sender_name": data.get("sender_name", "Vineet"),
        "status": data.get("status", "active"),
        "daily_limit": min(int(data.get("daily_limit", MAX_MESSAGES_PER_DAY)), MAX_MESSAGES_PER_DAY),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await _db.outreach_campaigns.insert_one(doc)
    return {"success": True, "campaign_id": campaign_id}


@router.put("/{campaign_id}")
async def update_campaign(campaign_id: str, data: dict):
    fields = {}
    for k in ["name", "message_template", "status", "daily_limit"]:
        if k in data:
            fields[k] = data[k]
    if not fields:
        raise HTTPException(status_code=400, detail="No fields to update")
    result = await _db.outreach_campaigns.update_one({"campaign_id": campaign_id}, {"$set": fields})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Campaign not found")
    return {"success": True}


@router.delete("/{campaign_id}")
async def delete_campaign(campaign_id: str):
    await _db.outreach_campaigns.delete_one({"campaign_id": campaign_id})
    await _db.campaign_prospects.delete_many({"campaign_id": campaign_id})
    return {"success": True}


# ============== Import ==============
@router.post("/{campaign_id}/import-prospects")
async def import_prospects_from_xlsx(campaign_id: str):
    campaign = await _db.outreach_campaigns.find_one({"campaign_id": campaign_id})
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")

    # Check multiple paths — local first, then cloud fallback
    xlsx_path = None
    for p in ["/tmp/prospects.xlsx", "/app/backend/uploads/prospects.xlsx"]:
        if os.path.exists(p):
            xlsx_path = p
            break
    if not xlsx_path:
        xlsx_path = _ensure_prospects_file()
    if not xlsx_path:
        raise HTTPException(status_code=400, detail="No prospects file found locally or in cloud storage.")

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    imported = 0
    skipped = 0

    for row in ws.iter_rows(min_row=5, values_only=True):
        if not row[0] or not str(row[0]).isdigit():
            continue
        linkedin_url = row[16] or ""
        if not linkedin_url or "linkedin.com" not in str(linkedin_url):
            skipped += 1
            continue
        public_id = str(linkedin_url).rstrip("/").split("/")[-1]
        existing = await _db.campaign_prospects.find_one({"campaign_id": campaign_id, "public_id": public_id})
        if existing:
            skipped += 1
            continue
        await _db.campaign_prospects.insert_one({
            "campaign_id": campaign_id, "rank": int(row[0]),
            "brand": row[1] or "", "name": row[17] or "", "title": row[21] or "",
            "category": row[2] or "", "pitch": row[13] or "",
            "linkedin_url": str(linkedin_url), "public_id": public_id,
            "status": "pending", "created_at": datetime.now(timezone.utc).isoformat(),
            "sent_at": None, "error": None,
        })
        imported += 1
    return {"success": True, "imported": imported, "skipped": skipped}


# ============== Prospects ==============
@router.get("/{campaign_id}/prospects")
async def get_campaign_prospects(campaign_id: str, status: str = None, limit: int = 100, skip: int = 0):
    query = {"campaign_id": campaign_id}
    if status:
        query["status"] = status
    prospects = []
    async for p in _db.campaign_prospects.find(query, {"_id": 0}).sort("rank", 1).skip(skip).limit(limit):
        prospects.append(p)
    total = await _db.campaign_prospects.count_documents(query)
    return {"prospects": prospects, "total": total}


@router.post("/{campaign_id}/retry-failed")
async def retry_failed_prospects(campaign_id: str):
    """Reset failed prospects back to pending for retry."""
    result = await _db.campaign_prospects.update_many(
        {"campaign_id": campaign_id, "status": "failed"},
        {"$set": {"status": "pending", "error": None, "sent_at": None}}
    )
    return {"success": True, "reset": result.modified_count}


# ============== Send (Background Task) ==============
@router.post("/{campaign_id}/send-batch")
async def trigger_send_batch(campaign_id: str, data: dict = None):
    """Trigger a batch send as a background task. Returns immediately."""
    if data is None:
        data = {}
    campaign = await _db.outreach_campaigns.find_one({"campaign_id": campaign_id})
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")

    # Pre-flight cookie check
    headers, err = await _validate_cookie()
    if err:
        raise HTTPException(status_code=400, detail=err)

    # Check daily limit
    today_start = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0).isoformat()
    sent_today = await _db.campaign_prospects.count_documents({
        "campaign_id": campaign_id, "status": "sent", "sent_at": {"$gte": today_start}
    })
    limit = campaign.get("daily_limit", MAX_MESSAGES_PER_DAY)
    remaining = limit - sent_today
    if remaining <= 0:
        return {"success": False, "detail": f"Daily limit reached ({limit}/day). Resume tomorrow.", "sent": 0}

    batch_size = min(data.get("batch_size", 10), remaining)
    pending = await _db.campaign_prospects.count_documents({"campaign_id": campaign_id, "status": "pending"})
    if pending == 0:
        return {"success": True, "detail": "No pending prospects left.", "sent": 0}

    # Launch background task
    asyncio.create_task(_send_batch_bg(campaign_id, campaign, headers, batch_size))
    return {"success": True, "detail": f"Sending batch of {min(batch_size, pending)} in background. Check status in a few minutes.", "batch_size": min(batch_size, pending), "daily_remaining": remaining}


async def _send_batch_bg(campaign_id: str, campaign: dict, headers: dict, batch_size: int):
    """Background task that sends messages with atomic claiming."""
    message_template = campaign.get("message_template", "")
    sent = 0
    failed = 0

    async with httpx.AsyncClient(timeout=30.0) as http_client:
        for _ in range(batch_size):
            # Atomic claim: find pending and set to 'sending'
            prospect = await _db.campaign_prospects.find_one_and_update(
                {"campaign_id": campaign_id, "status": "pending"},
                {"$set": {"status": "sending"}},
                sort=[("rank", 1)],
                return_document=True
            )
            if not prospect:
                break

            public_id = prospect["public_id"]
            try:
                # Personalize message
                first_name = (prospect.get("name", "").split()[0]) if prospect.get("name") else "there"
                message = message_template.replace("{name}", first_name).replace("{brand}", prospect.get("brand", "your brand"))

                # Resolve URN
                member_urn, err = await _resolve_member_urn(http_client, public_id, headers)
                if err == "cookie_expired":
                    # Cookie died — revert this and all remaining
                    await _db.campaign_prospects.update_one(
                        {"campaign_id": campaign_id, "public_id": public_id},
                        {"$set": {"status": "pending", "error": "Cookie expired — batch aborted"}}
                    )
                    logger.error(f"Campaign {campaign_id}: Cookie expired, aborting batch.")
                    break
                if not member_urn:
                    await _db.campaign_prospects.update_one(
                        {"campaign_id": campaign_id, "public_id": public_id},
                        {"$set": {"status": "failed", "error": err or "Could not resolve profile", "sent_at": datetime.now(timezone.utc).isoformat()}}
                    )
                    failed += 1
                    continue

                # Send message
                success, resp_text = await _send_single_message(http_client, headers, member_urn, public_id, message)
                if success:
                    await _db.campaign_prospects.update_one(
                        {"campaign_id": campaign_id, "public_id": public_id},
                        {"$set": {"status": "sent", "sent_at": datetime.now(timezone.utc).isoformat(), "error": None}}
                    )
                    sent += 1
                    logger.info(f"Campaign {campaign_id}: Sent to {prospect['name']} ({public_id})")
                else:
                    await _db.campaign_prospects.update_one(
                        {"campaign_id": campaign_id, "public_id": public_id},
                        {"$set": {"status": "failed", "error": resp_text[:200], "sent_at": datetime.now(timezone.utc).isoformat()}}
                    )
                    failed += 1

                # Rate limit
                await asyncio.sleep(8)

            except Exception as e:
                await _db.campaign_prospects.update_one(
                    {"campaign_id": campaign_id, "public_id": public_id},
                    {"$set": {"status": "failed", "error": str(e)[:200]}}
                )
                failed += 1

    logger.info(f"Campaign {campaign_id}: Batch done. Sent: {sent}, Failed: {failed}")


# ============== Auto Campaign Runner ==============
async def run_campaign_auto_sender():
    """Background task: sends campaign messages daily, respecting rate limits."""
    await asyncio.sleep(600)

    while True:
        try:
            active = await _db.outreach_campaigns.find({"status": "active"}).to_list(10)
            for campaign in active:
                cid = campaign["campaign_id"]
                pending = await _db.campaign_prospects.count_documents({"campaign_id": cid, "status": "pending"})
                if pending == 0:
                    continue

                today_start = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0).isoformat()
                sent_today = await _db.campaign_prospects.count_documents({
                    "campaign_id": cid, "status": "sent", "sent_at": {"$gte": today_start}
                })
                limit = campaign.get("daily_limit", MAX_MESSAGES_PER_DAY)
                if sent_today >= limit:
                    continue

                headers, err = await _validate_cookie()
                if err:
                    logger.warning(f"Campaign auto-sender: {err}")
                    break

                batch = min(5, limit - sent_today)
                logger.info(f"Campaign auto-sender: {cid} sending {batch} ({pending} pending, {sent_today}/{limit} today)")
                await _send_batch_bg(cid, campaign, headers, batch)

        except Exception as e:
            logger.error(f"Campaign auto-sender error: {e}")

        await asyncio.sleep(1800)
